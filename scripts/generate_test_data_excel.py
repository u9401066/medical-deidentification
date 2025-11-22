"""
Generate Test Medical Records
生成測試醫療病歷

Creates synthetic multi-language medical records for testing de-identification.
建立合成的多語言醫療病歷用於測試去識別化。

WARNING: All data is synthetic. Never use real PHI!
警告: 所有資料都是合成的。絕不使用真實個資！
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random


def generate_test_excel_multilanguage():
    """
    Generate complex multi-language Excel file with PHI
    生成包含個資的複雜多語言 Excel 檔案
    
    Languages: Traditional Chinese, English, Japanese, Korean, Thai
    語言: 繁體中文、英文、日文、韓文、泰文
    """
    wb = Workbook()
    
    # Sheet 1: Patient Demographics (多語言患者基本資料)
    ws1 = wb.active
    ws1.title = "Patient Demographics"
    
    # Headers with styling - added narrative description column
    headers = [
        "Patient ID\n病患編號\n患者ID",
        "Name\n姓名\n이름", 
        "DOB\n出生日期\n生年月日",
        "Age\n年齡\n나이",
        "Gender\n性別\n性別",
        "Phone\n電話\nโทรศัพท์",
        "Email\n電子郵件\n이메일",
        "Address\n地址\nที่อยู่",
        "Insurance\n保險號碼\n保険番号",
        "Patient Summary\n病患摘要\n患者概要"  # NEW: Narrative description with embedded PHI
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Generate 50 diverse patient records with narrative descriptions
    # Each tuple: (ID, Name, DOB, Age, Gender, Phone, Email, Address, Insurance, Narrative)
    patients = [
        # Traditional Chinese names with embedded PHI in narratives - EXTENDED TO >1024 TOKENS
        ("MRN-2024-001", "陳建國", "1955-03-15", 69, "男", "02-2234-5678", "chen.jg@hospital.tw", "台北市大安區信義路四段123號5樓", "NHI-A123456789",
         """【病歷摘要 - Medical Record Summary】
Patient Name: 陳建國 (Chen Jian-Guo) | MRN: MRN-2024-001 | DOB: 1955年3月15日 (March 15, 1955) | Current Age: 69歲 (69 years old) | Gender: 男性 (Male)
Contact Information: Mobile: 02-2234-5678 | Email: chen.jg@hospital.tw | Residence: 台北市大安區信義路四段123號5樓 (5F, No.123, Sec.4, Xinyi Rd., Da'an Dist., Taipei City)
Insurance: National Health Insurance Card No. NHI-A123456789

【Admission Details 入院資訊】
Admission Date: 2024年1月10日 (January 10, 2024, 10:30 AM)
Hospital: 國立台灣大學醫學院附設醫院心臟血管內科 (National Taiwan University Hospital, Department of Cardiovascular Medicine)
Ward: 心臟內科加護病房5A病房 (Cardiac ICU Ward 5A) | Bed Number: 512床 (Bed 512)
Attending Physician: 李文華主任醫師 (Dr. Lee Wen-Hua, Chief of Cardiology)
Admission Diagnosis: Acute Myocardial Infarction (急性心肌梗塞), Hypertension Stage 2 (第二期高血壓), Type 2 Diabetes Mellitus (第二型糖尿病)

【Chief Complaint 主訴】
The patient, Mr. 陳建國 (age 69, born March 15, 1955), presented to the emergency department at 09:45 AM on January 10, 2024, with severe substernal chest pain radiating to the left arm, accompanied by shortness of breath, diaphoresis, and nausea for the past 3 hours. Patient's wife called emergency services and mentioned calling 02-2234-5678 for medical history. Mr. Chen stated "I felt crushing chest pain while having breakfast at my home on 信義路 (Xinyi Road) around 7 AM this morning."

【Past Medical History 過去病史】
1. Hypertension - diagnosed 15 years ago (2009年確診，當時54歲) - poorly controlled despite medication
2. Type 2 Diabetes Mellitus - diagnosed 10 years ago (2014年確診，當時59歲) - HbA1c: 8.2%
3. Hyperlipidemia - total cholesterol 280 mg/dL, LDL 180 mg/dL
4. Former smoker - quit 5 years ago (2019年戒菸，當時64歲) - 30 pack-year history
5. Family history: Father died of MI at age 72 (父親於72歲死於心肌梗塞), mother has hypertension (母親有高血壓病史)

【Social History 社會史】
Patient 陳建國 (Chen Jian-Guo, DOB: 1955-03-15, currently 69 years old) is a retired high school teacher who taught mathematics for 35 years. He lives with his wife at 台北市大安區信義路四段123號5樓 (5th floor apartment on Xinyi Road Section 4). The apartment phone number is 02-2234-5678, which is also his mobile number. His wife can be reached at the same number. Email contact: chen.jg@hospital.tw for medical records or appointment scheduling. Mr. Chen mentioned he has been living at the Xinyi Road address since 1985 (民國74年), approximately 39 years ago when he was 30 years old. He walks to Da'an Park (大安森林公園) near his home on 信義路 three times per week for exercise.

【Physical Examination on Admission 入院理學檢查】
Date/Time: January 10, 2024, 10:45 AM | Location: Emergency Department, NTUH
Vital Signs: BP: 165/95 mmHg (elevated) | HR: 110 bpm (tachycardia) | RR: 24/min (tachypnea) | Temp: 37.2°C | SpO2: 92% on room air
General: Mr. 陳建國 (age 69) appeared anxious, diaphoretic, in moderate distress
Cardiovascular: Tachycardia, regular rhythm, no murmurs, JVP elevated at 8 cm
Respiratory: Bilateral basal crackles, consistent with pulmonary edema
Abdomen: Soft, non-tender, no organomegaly
Extremities: No peripheral edema, bilateral pedal pulses present

【Laboratory and Diagnostic Results 檢驗結果】
ECG (10:50 AM): ST-segment elevation in leads V1-V4, consistent with anterior wall STEMI
Cardiac Biomarkers: Troponin-I: 5.8 ng/mL (markedly elevated, normal <0.04), CK-MB: 156 U/L (elevated)
Complete Blood Count: WBC 12,500/μL, Hb 13.8 g/dL, Platelets 245,000/μL
Chemistry: Glucose 245 mg/dL, BUN 28 mg/dL, Creatinine 1.3 mg/dL, eGFR 55 mL/min
Lipid Panel: Total Cholesterol 280 mg/dL, LDL 180 mg/dL, HDL 35 mg/dL, TG 325 mg/dL
Chest X-ray: Cardiomegaly, pulmonary congestion

【Treatment Course 治療經過】
Patient 陳建國 (MRN-2024-001, born 1955-03-15, age 69, insurance NHI-A123456789) underwent emergency cardiac catheterization at 12:30 PM on January 10, 2024. Procedure revealed 95% stenosis of left anterior descending (LAD) artery. Successful percutaneous coronary intervention (PCI) with drug-eluting stent placement was performed by Dr. 李文華. Patient was transferred to Cardiac ICU Ward 5A, Bed 512 at 15:00 hours. Contact number 02-2234-5678 was called to notify family. Post-procedure, patient received dual antiplatelet therapy (Aspirin 100mg + Clopidogrel 75mg), beta-blocker (Metoprolol 50mg BID), ACE inhibitor (Ramipril 5mg daily), statin (Atorvastatin 80mg daily), and insulin therapy for glucose control.

【Progress Notes 病程記錄】
Day 1 (Jan 10, 2024): Mr. Chen (陳建國, 69歲) stable post-PCI, chest pain resolved, vitals stable in Ward 5A Bed 512
Day 2 (Jan 11, 2024): Patient 陳建國 ambulating with assistance, no recurrent chest pain, troponin trending down
Day 3 (Jan 12, 2024): Echocardiogram showed LVEF 45%, anterior wall hypokinesis, patient教育 regarding medication compliance
Day 4 (Jan 13, 2024): Patient ready for discharge, follow-up arranged at cardiac clinic

【Discharge Summary 出院摘要】
Discharge Date: January 13, 2024, 10:00 AM
Patient: 陳建國 (Chen Jian-Guo), Male, Age 69 (Born: March 15, 1955), MRN: MRN-2024-001
Discharge Address: 台北市大安區信義路四段123號5樓 | Contact: 02-2234-5678 | Email: chen.jg@hospital.tw
Insurance: NHI-A123456789
Discharge Diagnosis: (1) Acute Anterior Wall ST-Elevation Myocardial Infarction, status post PCI with stent to LAD (2) Hypertension (3) Type 2 Diabetes Mellitus (4) Hyperlipidemia
Discharge Medications: Aspirin, Clopidogrel, Metoprolol, Ramipril, Atorvastatin, Insulin Glargine
Follow-up: Cardiac Rehabilitation Program starting January 20, 2024 | Cardiology Clinic appointment February 10, 2024, 2:00 PM with Dr. 李文華
Emergency Contact: Please call 02-2234-5678 or email chen.jg@hospital.tw for any concerns. Patient 陳建國 (age 69, DOB 1955-03-15) residing at 信義路四段123號5樓 should seek immediate care if chest pain recurs.

【臨床醫師簽名】Dr. 李文華, Chief of Cardiology, NTUH | Date: January 13, 2024"""),
        
        ("MRN-2024-002", "林雅婷", "1992-07-22", 32, "女", "04-2345-6789", "lin.yt@email.com", "台中市西屯區文心路二段456號", "NHI-B987654321",
         """【Obstetrics & Gynecology Medical Record 婦產科病歷】
Patient: 林雅婷 (Lin Ya-Ting) | MRN: MRN-2024-002 | DOB: 1992年7月22日 (July 22, 1992) | Age: 32歲 (32 years old) | Gender: 女性 (Female)
Contact: Mobile 04-2345-6789 | Email: lin.yt@email.com | Address: 台中市西屯區文心路二段456號 (No.456, Sec.2, Wenxin Rd., Xitun Dist., Taichung City)
National Health Insurance Number: NHI-B987654321

【Admission Information 住院資訊】
Admission Date/Time: January 15, 2024, 08:30 AM
Hospital: 台中榮民總醫院婦產部 (Taichung Veterans General Hospital, Department of Obstetrics & Gynecology)
Ward/Bed: 婦產科病房3B, Bed 305 (OB/GYN Ward 3B, Bed Number 305)
Attending Physician: 張美玲主治醫師 (Dr. Chang Mei-Ling, Attending Obstetrician)
Admission Diagnosis: Pregnancy at 38 weeks gestation, Gestational Diabetes Mellitus, Previous Cesarean Section

【Chief Complaint 主訴】
Ms. 林雅婷 (Lin Ya-Ting, 32-year-old female, born July 22, 1992) presented to labor and delivery at 08:15 AM on January 15, 2024, with regular uterine contractions occurring every 5 minutes for the past 2 hours. Patient called from her residence at 台中市西屯區文心路二段456號 using mobile phone 04-2345-6789 before arriving. She mentioned feeling contractions starting around 6 AM while preparing breakfast at her home on Wenxin Road (文心路). This is her second pregnancy. Patient email lin.yt@email.com is on file for appointment reminders and test results.

【Obstetric History 產科病史】
Gravida 2, Para 1 (G2P1)
Previous Pregnancy (2020): 
- First child delivered via Cesarean Section on March 15, 2020, at 台中榮總 when patient 林雅婷 was 27 years old (born 1992-07-22)
- Indication for C/S: Cephalopelvic disproportion (CPD)
- Male infant, birth weight 3,580 grams, healthy, now 4 years old
- Patient recovered well post-surgery, no complications
- Patient mentioned she was living at the same address on 文心路二段456號 during first pregnancy

Current Pregnancy (2024):
- Last Menstrual Period (LMP): April 20, 2023
- Expected Date of Delivery (EDD): January 25, 2024
- Gestational Age: 38 weeks + 3 days
- Patient 林雅婷 (age 32, DOB 1992/07/22, insurance NHI-B987654321) attended regular prenatal visits every month
- All prenatal care received at 台中榮總婦產科 with Dr. 張美玲
- Contact number 04-2345-6789 used for all appointment confirmations

【Prenatal Course 產前追蹤】
First Trimester (Weeks 1-12):
- Initial prenatal visit: May 25, 2023, patient 林雅婷 confirmed pregnancy via urine hCG and ultrasound
- Dating ultrasound: Single viable intrauterine pregnancy, CRL consistent with dates
- No morning sickness, patient continued working as accountant at firm near her home on 文心路

Second Trimester (Weeks 13-28):
- Anatomy scan at 20 weeks (September 2023): Normal fetal anatomy, placenta anteriorly located
- Glucose Challenge Test (GCT) at 24 weeks: 165 mg/dL (abnormal, threshold >140 mg/dL)
- Oral Glucose Tolerance Test (OGTT) at 25 weeks: Fasting 98 mg/dL, 1-hr 190 mg/dL, 2-hr 170 mg/dL, 3-hr 145 mg/dL
- Diagnosis: Gestational Diabetes Mellitus (GDM) confirmed on October 15, 2023
- Patient 林雅婷 (then 31 years old, now 32) referred to dietitian for meal planning
- Blood glucose monitoring initiated, patient using home glucometer, results sent via email lin.yt@email.com

Third Trimester (Weeks 29-40):
- Patient 林雅婷 maintained good glycemic control with diet modification
- Fasting glucose: 85-95 mg/dL, Postprandial glucose: 110-130 mg/dL
- Fetal growth scan at 32 weeks: Estimated fetal weight 2,100 grams (appropriate for gestational age)
- Fetal growth scan at 36 weeks: Estimated fetal weight 2,850 grams, amniotic fluid normal, biophysical profile 8/8
- Non-stress tests weekly since 36 weeks: All reactive
- Patient contact 04-2345-6789 called for NST scheduling

【Past Medical History 病史】
1. Gestational Diabetes Mellitus - current pregnancy (diagnosed October 2023 at age 31)
2. Previous Cesarean Section - March 2020 (age 27) for CPD
3. Allergic Rhinitis - since childhood (diagnosed around age 8 in year 2000)
4. No chronic medical conditions, no hypertension, no cardiac disease
5. Family History: Mother (林母, age 58) has Type 2 Diabetes diagnosed at age 50, Father (林父, age 61) healthy

【Medications 用藥記錄】
Current Medications for patient 林雅婷 (MRN-2024-002, DOB 1992-07-22):
1. Prenatal vitamins with iron (since May 2023)
2. Folic Acid 400 mcg daily (since May 2023)
3. Calcium supplement 500 mg BID (since June 2023)
4. No insulin required for GDM management (diet-controlled)

Allergies: No Known Drug Allergies (NKDA) - patient 林雅婷 confirmed at admission on Jan 15, 2024

【Physical Examination 身體檢查】
Admission Date: January 15, 2024, 08:45 AM | Location: Labor & Delivery Unit, 台中榮總
Patient: 林雅婷 (Lin Ya-Ting, female, 32 years old, born July 22, 1992)

Vital Signs: BP 125/78 mmHg | HR 88 bpm | RR 18/min | Temp 36.8°C | SpO2 98% on room air
General Appearance: Ms. 林 alert, oriented, in active labor, contractions every 4-5 minutes
HEENT: Normocephalic, atraumatic, moist mucous membranes
Cardiovascular: Regular rate and rhythm (RRR), no murmurs
Respiratory: Clear breath sounds bilaterally, no respiratory distress
Abdomen: Gravid abdomen, fundal height 36 cm, vertex presentation, fetal heart rate 140 bpm (reactive, Category I)
Extremities: Trace bilateral lower extremity edema, no calf tenderness

Obstetric Examination:
- Cervical dilation: 4 cm | Effacement: 70% | Station: -1 | Presenting part: Vertex (頭位)
- Membranes: Intact | Contraction pattern: Regular, every 4-5 minutes, lasting 50-60 seconds, moderate intensity
- Fetal Heart Rate Monitoring: Baseline 140 bpm, moderate variability, accelerations present, no decelerations (Category I tracing)

【Labor Course & Delivery 產程與生產】
January 15, 2024:
09:30 AM - Patient 林雅婷 (age 32, bed 305 in Ward 3B) cervix 5 cm dilated, contractions q3-4 min
10:45 AM - Cervix 7 cm, station 0, patient requesting epidural analgesia, anesthesia consult called
11:15 AM - Epidural placed by Dr. 王麻醉師 without complications, patient comfortable, vital signs stable
12:30 PM - Cervix fully dilated (10 cm), station +1, patient感覺有便意感 (feeling urge to push)
12:45 PM - Decision made for repeat Cesarean Section due to previous C/S and concern for uterine rupture risk
         Patient 林雅婷 (MRN-2024-002, insurance NHI-B987654321) consented for procedure
         Emergency contact 04-2345-6789 called to notify family
01:15 PM - Patient transferred to Operating Room 3 for repeat Cesarean Section
01:45 PM - Cesarean delivery performed by Dr. 張美玲 with Dr. 李副手醫師 assisting
         Low transverse uterine incision, delivered healthy female infant
         Time of birth: 13:52 PM (1:52 PM) on January 15, 2024
         Infant weight: 3,250 grams | Length: 50 cm | Head circumference: 34 cm
         Apgar scores: 9 at 1 minute, 10 at 5 minutes
         Infant cried immediately, pink, active, no resuscitation required
02:15 PM - Surgery completed, estimated blood loss 600 mL, patient stable
02:45 PM - Patient transferred to Recovery Room, vital signs stable
04:00 PM - Patient 林雅婷 transferred back to Ward 3B Bed 305, alert and comfortable

【Postoperative Course 術後病程】
Day 0 (Jan 15, 2024, evening):
Patient 林雅婷 (32歲, DOB 1992-07-22) stable post-op in Ward 3B Bed 305. Pain controlled with IV PCA (Patient-Controlled Analgesia). Vitals stable. Foley catheter in place, output adequate. Fundus firm at umbilicus. Incision clean, dry, intact. Infant doing well in normal newborn nursery. Family visited, contact via 04-2345-6789.

Day 1 (Jan 16, 2024):
Patient ambulating with assistance. Foley catheter removed at 6 AM, voiding spontaneously. Tolerating clear liquid diet. Pain score 3-4/10 with oral analgesics. Breastfeeding initiated with lactation consultant support. Infant rooming-in with mother. Ms. 林雅婷 reports gas pain, bowel sounds present, passing flatus. Vital signs stable, afebrile. Incision checked - clean, no signs of infection.

Day 2 (Jan 17, 2024):
Patient 林雅婷 (MRN-2024-002, address 文心路二段456號) tolerating regular diet. Ambulating independently. Breastfeeding well established. Infant feeding every 2-3 hours. Pain well controlled with oral Ibuprofen 400mg q6h PRN. Staples to incision intact. Plan for discharge tomorrow if continues to progress well. Discharge teaching initiated regarding wound care, signs of infection, breastfeeding, postpartum depression screening.

Day 3 (Jan 18, 2024):
Ready for discharge. Patient 林雅婷 (age 32, insurance NHI-B987654321) ambulating without difficulty, pain controlled, incision healing well. Infant passed newborn hearing screen and metabolic screen. Jaundice level within normal limits. Both mother and baby cleared for discharge.

【Discharge Summary 出院摘要】
Discharge Date: January 18, 2024, 11:00 AM
Patient: 林雅婷 (Lin Ya-Ting), Female, 32 years old (Born: July 22, 1992), MRN: MRN-2024-002
Discharge Address: 台中市西屯區文心路二段456號 (No.456, Sec.2, Wenxin Rd., Xitun Dist., Taichung)
Contact: Mobile 04-2345-6789 | Email: lin.yt@email.com | Insurance: NHI-B987654321

Hospital Course: Patient admitted January 15, 2024, in active labor at 38+3 weeks gestation. Underwent repeat Cesarean Section. Delivered healthy female infant weighing 3,250g. Postoperative recovery uncomplicated.

Discharge Diagnosis:
1. Status post Repeat Cesarean Section, January 15, 2024
2. Delivery of viable female infant at 38+3 weeks gestation
3. Gestational Diabetes Mellitus (diet-controlled)
4. Previous Cesarean Section (2020)

Discharge Medications:
1. Ibuprofen 400mg PO q6h PRN for pain
2. Colace 100mg PO BID for constipation prevention
3. Prenatal vitamins one tablet daily
4. Iron supplement 325mg daily

Discharge Instructions:
- Wound care: Keep incision clean and dry, remove staples in 10-14 days at OB clinic
- Activity: Light activity only, no heavy lifting >10 pounds for 6 weeks, no driving for 2 weeks
- Diet: Regular diet, increase fiber and fluids
- Follow-up: Postpartum visit with Dr. 張美玲 on February 15, 2024 (4 weeks post-delivery)
- Pediatric follow-up for infant scheduled February 1, 2024
- Call 04-2345-6789 or email lin.yt@email.com if any concerns
- Return to hospital immediately if: fever >38°C, heavy vaginal bleeding, severe abdominal pain, wound drainage/redness, chest pain, shortness of breath

Patient 林雅婷 (age 32, born 1992-07-22, residing at 文心路二段456號) understands all discharge instructions and has no questions. Emergency contact: 04-2345-6789.

【主治醫師】Dr. 張美玲, OB/GYN, Taichung Veterans General Hospital | Date: January 18, 2024"""),
        
        ("MRN-2024-003", "王大明", "1978-11-08", 46, "男", "07-3456-7890", "wang.dm@example.tw", "高雄市苓雅區中山二路789號12樓", "NHI-C456789123",
         """【Orthopedic Surgery Medical Record 骨科病歷】
Patient Name: 王大明 (Wang Da-Ming) | MRN: MRN-2024-003 | DOB: 1978年11月8日 (November 8, 1978) | Age: 46歲 (46 years old) | Gender: 男性 (Male)
Contact Information: Phone: 07-3456-7890 | Email: wang.dm@example.tw | Address: 高雄市苓雅區中山二路789號12樓 (12F, No.789, Zhongshan 2nd Rd., Lingya Dist., Kaohsiung City)
Insurance: National Health Insurance Card Number: NHI-C456789123

【Emergency Department Visit 急診就醫】
Arrival Date/Time: February 1, 2024, 22:45 PM (10:45 PM)
Hospital: 高雄長庚紀念醫院急診部 (Chang Gung Memorial Hospital Kaohsiung, Emergency Department)
Mode of Arrival: Ambulance (救護車) from accident scene
Chief Complaint: Right lower extremity pain following motorcycle accident

【History of Present Illness 現病史】
Mr. 王大明 (Wang Da-Ming, 46-year-old male, born November 8, 1978) was riding his motorcycle near his apartment building at 高雄市苓雅區中山二路789號 around 22:00 (10 PM) on February 1, 2024, when he was struck by a car making an illegal left turn. Patient recalls the accident clearly - he was riding home from work to his 12th floor apartment on Zhongshan 2nd Road (中山二路789號12樓) when the collision occurred at the intersection of Zhongshan Road and Chenggong Road (中山路與成功路口). Emergency services were called by bystander. Paramedics arrived at scene within 8 minutes. Patient was conscious and alert but complaining of severe right leg pain and inability to bear weight. Patient's wallet containing insurance card NHI-C456789123 and ID showing address 中山二路789號12樓 was retrieved from accident scene. Emergency contact number 07-3456-7890 was called from patient's mobile phone to notify family.

Mechanism of Injury: Motor vehicle versus motorcycle collision at approximately 40-50 km/h. Patient 王大明 (age 46, DOB 1978-11-08) was wearing helmet (安全帽). No loss of consciousness. Patient struck on right lateral side, thrown from motorcycle, landed on right side. Airbag not applicable. Seatbelt not applicable (motorcycle).

【Emergency Department Assessment 急診評估】
Arrival Time: February 1, 2024, 22:45 PM
Primary Survey (ATLS Protocol):
A (Airway): Patent, patient speaking in full sentences, cervical collar applied for precaution
B (Breathing): Respiratory rate 20/min, SpO2 97% on room air, breath sounds equal bilaterally, no pneumothorax
C (Circulation): Blood pressure 138/85 mmHg, heart rate 105 bpm (tachycardia - likely pain-related), capillary refill <2 seconds, no active bleeding
D (Disability): Glasgow Coma Scale 15/15 (E4V5M6), pupils equal round reactive to light (PERRL), moves all extremities except right lower extremity due to pain
E (Exposure): Multiple abrasions on right side, obvious deformity of right lower leg

Vital Signs on Arrival:
BP: 138/85 mmHg | HR: 105 bpm | RR: 20/min | Temp: 36.9°C | SpO2: 97% on RA | Pain Score: 9/10 (right leg)

Secondary Survey:
Head: Helmet removed carefully, no scalp laceration, no Battle sign, no raccoon eyes
Face: No facial fractures, no dental trauma
Neck: C-collar in place, no midline tenderness (cleared clinically and with CT)
Chest: No chest wall tenderness, no crepitus, equal breath sounds, no rib fractures on exam
Abdomen: Soft, non-tender, no guarding or rigidity, no seat belt sign (N/A for motorcycle)
Pelvis: Stable to compression, no tenderness
Right Lower Extremity: 
  - Obvious deformity mid-shaft tibia and fibula
  - Swelling, ecchymosis
  - Skin intact, no open fracture
  - Dorsalis pedis pulse palpable, posterior tibial pulse palpable (neurovascular intact)
  - Sensation intact to light touch in superficial peroneal, deep peroneal, sural, saphenous, tibial nerve distributions
  - Unable to assess motor function due to pain and suspected fracture
  - Compartments soft, no signs of compartment syndrome on initial exam
Left Lower Extremity: Normal, no deformity, full range of motion
Upper Extremities: Bilateral abrasions, no fractures, full range of motion, neurovascularly intact

【Emergency Department Management 急診處置】
Time: February 1, 2024, 23:00-23:45 PM
1. IV access established - two 18G peripheral IVs placed in bilateral antecubital fossae
2. Normal saline 500 mL bolus administered
3. Pain management: Morphine 4mg IV push, later additional 2mg IV (total 6mg), pain improved to 6/10
4. Right lower extremity splinted in long leg posterior splint for immobilization
5. Tetanus toxoid 0.5 mL IM administered (last tetanus shot >10 years ago per patient 王大明 report)
6. NPO (nothing by mouth) status initiated in preparation for possible surgery
7. Labs drawn: CBC, CMP, coagulation studies
8. Imaging ordered: Right leg X-rays (AP and lateral), Chest X-ray (trauma protocol), Pelvis X-ray, Cervical spine CT

【Laboratory Results 檢驗報告】
Blood Draw Time: February 1, 2024, 23:15 PM
Complete Blood Count (CBC):
- WBC: 11,200/μL (slightly elevated, stress response)
- Hemoglobin: 14.5 g/dL (normal)
- Hematocrit: 43% (normal)
- Platelets: 265,000/μL (normal)

Comprehensive Metabolic Panel (CMP):
- Sodium: 140 mEq/L | Potassium: 4.1 mEq/L | Chloride: 102 mEq/L | Bicarbonate: 24 mEq/L
- BUN: 18 mg/dL | Creatinine: 0.9 mg/dL (eGFR >90 mL/min) | Glucose: 128 mg/dL (stress hyperglycemia)
- Calcium: 9.2 mg/dL | Total Protein: 7.0 g/dL | Albumin: 4.2 g/dL
- AST: 28 U/L | ALT: 32 U/L | Alkaline Phosphatase: 78 U/L | Total Bilirubin: 0.8 mg/dL

Coagulation Studies:
- PT: 12.5 seconds (normal) | INR: 1.0 | PTT: 28 seconds (normal)

Urinalysis: Clear, no hematuria

Blood Type: O positive

【Imaging Studies 影像學檢查】
1. Right Lower Extremity X-rays (February 1, 2024, 23:30 PM):
   AP and Lateral views of right tibia and fibula
   Findings: 
   - Closed, displaced, comminuted fracture of right tibial shaft at mid-diaphysis (middle third)
   - Associated fibular fracture at similar level
   - Fracture pattern: Oblique, comminuted (粉碎性骨折) with 3 main fragments
   - Displacement: Approximately 1 cm of shortening, 10-degree angulation
   - No intra-articular extension
   - No open fracture (skin intact on physical exam)
   
2. Chest X-ray (Portable AP): 
   No pneumothorax, no hemothorax, no rib fractures, cardiomediastinal silhouette normal
   
3. Pelvis X-ray (AP):
   No pelvic fractures, hip joints normal

4. Cervical Spine CT:
   No fractures, no subluxation, cervical collar cleared

【Orthopedic Consultation 骨科會診】
Consultation Time: February 2, 2024, 00:30 AM (12:30 AM)
Consulting Physician: 林建宏主治醫師 (Dr. Lin Jian-Hong), Orthopedic Trauma Surgeon

Assessment: Patient 王大明 (MRN-2024-003, DOB 1978-11-08, age 46 years, insurance NHI-C456789123, contact 07-3456-7890, address 高雄市苓雅區中山二路789號12樓) with closed, displaced, comminuted fracture of right tibial and fibular shafts following motorcycle versus car collision. Neurovascularly intact distal extremity. No compartment syndrome. No other significant injuries.

Recommendation: Operative fixation with intramedullary nailing (骨髓內釘固定術). Patient counseled regarding surgical procedure, risks, benefits, alternatives. Consent obtained for surgery. Family contacted at 07-3456-7890.

【Admission & Preoperative Preparation 入院及術前準備】
Admission Time: February 2, 2024, 01:00 AM
Ward: 骨科病房7A (Orthopedic Ward 7A) | Bed: 718床 (Bed 718)
Admitting Diagnosis: Closed right tibial and fibular shaft fractures
Surgery Scheduled: February 2, 2024, 08:00 AM - Intramedullary nailing of right tibia

Preoperative Orders for patient 王大明 (age 46, born Nov 8, 1978):
- NPO after midnight (maintained since ED arrival)
- IV fluids: Normal saline at 100 mL/hr
- Pain management: Morphine PCA (Patient-Controlled Analgesia) pump
- DVT prophylaxis: Sequential compression devices (SCDs) to bilateral lower extremities
- Antibiotics: Cefazolin 2g IV on call to OR (surgical prophylaxis)
- Preoperative labs: Reviewed and acceptable for surgery
- Consent: Signed by patient, witnessed by nurse, copy in chart
- Anesthesia evaluation: Completed by Dr. 陳麻醉科醫師, ASA Class II, cleared for general anesthesia
- Patient contact information confirmed: 07-3456-7890, wang.dm@example.tw, address 中山二路789號12樓

【Surgical Procedure 手術記錄】
Date of Surgery: February 2, 2024
Procedure Start Time: 08:15 AM | Procedure End Time: 10:45 AM | Total Duration: 2 hours 30 minutes
Operating Room: OR 5, 高雄長庚紀念醫院
Surgeon: Dr. 林建宏 (Lin Jian-Hong), Orthopedic Trauma Surgeon
Assistant Surgeon: Dr. 黃副手醫師 (Dr. Huang)
Anesthesiologist: Dr. 陳麻醉科醫師 (Dr. Chen), General Anesthesia
Scrub Nurse: 護理師王小姐 (Nurse Wang)

Preoperative Diagnosis: Closed, displaced, comminuted fracture of right tibial shaft; Associated right fibular shaft fracture
Postoperative Diagnosis: Same

Procedure Performed: 
1. Closed reduction and intramedullary nailing of right tibia (右側脛骨閉合復位骨髓內釘固定術)
2. Application of proximal and distal interlocking screws

Operative Findings:
Patient 王大明 (MRN-2024-003, age 46, DOB 1978-11-08) positioned supine on radiolucent operating table. General anesthesia induced without complications. Right lower extremity prepped and draped in sterile fashion. Fluoroscopic imaging confirmed fracture pattern: comminuted mid-shaft tibial fracture with 3 fragments and associated fibular fracture. Neurovascular exam stable preoperatively.

Operative Technique (Detailed):
1. Small incision (approximately 3 cm) made just medial to tibial tuberosity at proximal tibia
2. Entry point established at medial aspect of tibial plateau under fluoroscopic guidance
3. Guidewire inserted across fracture site under fluoroscopic visualization
4. Fracture reduced using manual traction and manipulation
5. Reaming performed sequentially from 8mm to 11mm diameter
6. Intramedullary nail (11mm x 340mm titanium nail) inserted over guidewire under fluoroscopy
7. Proximal interlocking: Two 4.5mm screws placed from medial to lateral through nail
8. Distal interlocking: Two 4.5mm screws placed from lateral to medial through nail
9. Fluoroscopic images confirmed excellent fracture reduction and hardware placement
10. Wound irrigated copiously with normal saline
11. Fascia, subcutaneous tissue, and skin closed in layers with absorbable sutures
12. Sterile dressing applied, well-padded posterior splint applied for additional support

Estimated Blood Loss: 150 mL
Intravenous Fluids: 1,500 mL Lactated Ringer's solution
Urine Output: 400 mL (Foley catheter placed after anesthesia induction)
Specimens: None sent to pathology
Implants: 11mm x 340mm titanium intramedullary tibial nail, four 4.5mm interlocking screws
Complications: None
Patient tolerated procedure well, transferred to PACU (Post-Anesthesia Care Unit) in stable condition

【Postoperative Course 術後病程】
Post-Op Day 0 (February 2, 2024, afternoon):
Patient 王大明 (age 46, Ward 7A Bed 718, contact 07-3456-7890) recovered from anesthesia without complications. Alert, oriented x3. Transferred from PACU to Orthopedic Ward 7A Bed 718 at 13:00 (1 PM). Vital signs stable: BP 128/75, HR 78, Temp 36.8°C. Pain controlled with PCA morphine. Right lower extremity neurovascularly intact - dorsalis pedis pulse palpable, able to wiggle toes, sensation intact. Posterior splint in place. Ice applied to reduce swelling. DVT prophylaxis continued with SCDs and started subcutaneous Enoxaparin 40mg daily. IV antibiotics: Cefazolin 1g q8h x 24 hours. NPO initially, advanced to clear liquids in evening, tolerating well. Foley catheter draining clear yellow urine. Family visited, updated on surgery outcome, contact info 07-3456-7890 on file.

Post-Op Day 1 (February 3, 2024):
Patient 王大明 (MRN-2024-003, born 1978-11-08, address 中山二路789號12樓, insurance NHI-C456789123) reporting pain 4-5/10, controlled with oral oxycodone 5mg q4h PRN. Transitioned off PCA. Foley catheter removed at 6 AM, voiding spontaneously without difficulty. Diet advanced to regular, tolerating well. Physical therapy consulted - patient educated on non-weight bearing status for right lower extremity, crutch walking training initiated. Patient able to ambulate 50 feet with crutches and supervision. Right leg neurovascularly intact, no signs of compartment syndrome. Incision clean, dry, intact. Splint in place. Temperature 37.1°C (afebrile). Plan: Continue current management, advance activity as tolerated.

Post-Op Day 2 (February 4, 2024):
Patient 王大明 (46歲) ambulating independently with crutches, non-weight bearing on right leg. Pain well controlled with oral medications. Incision healing well, no signs of infection. X-rays obtained: hardware in good position, fracture alignment maintained. Physical therapy progressing well - patient independent with crutches, able to manage stairs. Discharge planning initiated. Social work consulted to arrange home modifications and equipment (shower chair, raised toilet seat, etc.) at patient's apartment at 中山二路789號12樓 (12th floor - elevator access confirmed). Patient王大明 lives with wife and two children (son age 12, daughter age 9) who can assist with activities of daily living.

Post-Op Day 3 (February 5, 2024):
Ready for discharge. Patient 王大明 (contact 07-3456-7890, email wang.dm@example.tw) meeting all discharge criteria: pain controlled on oral medications, ambulatory with crutches non-weight bearing, neurovascularly intact, no signs of infection or complications, patient and family understand discharge instructions.

【Discharge Summary 出院摘要】
Discharge Date: February 5, 2024, 10:00 AM
Patient: 王大明 (Wang Da-Ming), Male, 46 years old (Born: November 8, 1978), MRN: MRN-2024-003
Discharge Address: 高雄市苓雅區中山二路789號12樓 (12F, No.789, Zhongshan 2nd Rd., Lingya Dist., Kaohsiung)
Contact: Phone 07-3456-7890 | Email: wang.dm@example.tw | Insurance: NHI-C456789123

Admission Date: February 2, 2024 | Discharge Date: February 5, 2024 | Length of Stay: 4 days

Hospital Course: Patient 王大明 (age 46, DOB 1978-11-08) admitted following motorcycle accident on February 1, 2024, near his residence at 中山二路789號12樓. Sustained closed, displaced, comminuted fractures of right tibial and fibular shafts. Underwent successful intramedullary nailing of right tibia on February 2, 2024. Postoperative course uncomplicated. Pain controlled, neurovascularly intact, ambulating with crutches. Ready for discharge home.

Discharge Diagnosis:
1. Status post intramedullary nailing of right tibia, February 2, 2024
2. Closed, comminuted fracture of right tibial shaft
3. Closed fracture of right fibular shaft
4. Status post motorcycle versus automobile collision, February 1, 2024

Discharge Medications:
1. Oxycodone 5mg PO q4-6h PRN for moderate to severe pain (#30 tablets, no refills - narcotic contract signed)
2. Acetaminophen 500mg PO q6h PRN for mild pain
3. Ibuprofen 400mg PO TID with food for inflammation (take with food to prevent GI upset)
4. Enoxaparin (Lovenox) 40mg subcutaneous daily x 4 weeks for DVT prophylaxis (patient/family taught injection technique)
5. Docusate 100mg PO BID for constipation prevention (related to narcotics)
6. Calcium 500mg + Vitamin D 400 IU daily for bone healing

Discharge Instructions for Mr. 王大明 (age 46, born Nov 8, 1978, contact 07-3456-7890):
1. Weight-Bearing Status: NON-WEIGHT BEARING on right lower extremity for 6 weeks (absolutely no weight on right leg)
2. Ambulation: Use crutches at all times when walking, continue as taught by physical therapy
3. Wound Care: Keep surgical incision clean and dry, may shower after 48 hours (cover with plastic bag), no soaking in bathtub/pool for 2 weeks
4. Activity: Elevate right leg above heart level when sitting or lying down to reduce swelling, ice pack 20 minutes q2-3h for first week
5. Diet: Regular diet, high protein and calcium for bone healing, increase fluids and fiber due to narcotic constipation risk
6. DVT Prevention: Continue Enoxaparin injections daily, move ankle up and down frequently, no prolonged sitting
7. Pain Management: Take pain medications as prescribed, do not drive while taking narcotics

Follow-up Appointments:
1. Wound check in Orthopedic Clinic: February 12, 2024 (7 days post-op) at 10:00 AM with Dr. 林建宏
2. X-ray and suture removal: February 19, 2024 (2 weeks post-op) at 2:00 PM
3. Follow-up X-ray: March 19, 2024 (6 weeks post-op) - will assess fracture healing and advance weight-bearing
4. Call clinic at 07-7317123 to schedule or reschedule appointments

Warning Signs - Call 07-3456-7890 or return to Emergency Department immediately if:
- Fever >38.5°C (101.3°F)
- Increased pain not relieved by medications
- Numbness, tingling, or inability to move toes
- Cold, pale, or blue discoloration of right foot
- Drainage, redness, or foul odor from surgical incision
- Calf pain, swelling, or warmth (signs of blood clot)
- Chest pain or shortness of breath (signs of pulmonary embolism)
- Excessive bleeding from incision

Return to Work: Patient王大明 works as office manager (sedentary job) - may return to work in 2-3 weeks if pain controlled and can manage transportation. No driving until off narcotics and cleared by physician (approximately 4-6 weeks).

Disability Paperwork: Provided for employer, patient expected to be off work for 4-6 weeks minimum, full recovery 3-4 months.

Patient 王大明 (MRN-2024-003, age 46, address 高雄市苓雅區中山二路789號12樓, phone 07-3456-7890, email wang.dm@example.tw, insurance NHI-C456789123) verbalizes understanding of all discharge instructions, has no questions at this time. Wife present during discharge teaching and will assist at home.

【主治醫師】Dr. 林建宏 (Lin Jian-Hong), Orthopedic Trauma Surgery, 高雄長庚紀念醫院 | Date: February 5, 2024"""),
        
        # English names with embedded PHI - EXTENDED TO >1024 TOKENS
        ("MRN-2024-004", "John Smith", "1963-05-12", 61, "Male", "+886-2-8765-4321", "john.smith@hospital.org", "15F, No. 234, Renai Road, Taipei", "INS-US-789456123",
         """【Internal Medicine - Endocrinology Service 內科內分泌科病歷】
Patient: John Michael Smith | Medical Record Number: MRN-2024-004 | Date of Birth: May 12, 1963 | Current Age: 61 years old | Sex: Male
Contact: Mobile +886-2-8765-4321 | Email: john.smith@hospital.org | Residential Address: 15F, No. 234, Renai Road, Section 3, Da'an District, Taipei City 106, Taiwan
Insurance Provider: International Insurance - Policy Number INS-US-789456123 | Nationality: United States Citizen, Permanent Resident of Taiwan

【Chief Complaint & Admission Information】
Date of Admission: March 5, 2024, at 2:30 PM
Hospital: Taipei Veterans General Hospital (台北榮民總醫院), Endocrinology Department
Ward/Room: Endocrinology Ward 6B, Room 605, Bed A
Attending Physician: Dr. Michael Chen (陳明哲醫師), MD, PhD, Endocrinology Specialist
Reason for Admission: Uncontrolled Type 2 Diabetes Mellitus with hyperglycemia (blood glucose >400 mg/dL), diabetic ketoacidosis risk

Chief Complaint: Mr. John Smith, a 61-year-old American expatriate (born May 12, 1963 in New York City, USA), presented to the outpatient endocrinology clinic on March 5, 2024, with complaints of polyuria (excessive urination), polydipsia (excessive thirst), unexplained weight loss of 8 kilograms over the past 3 months, and fatigue for the past 2 weeks. Patient called clinic from his apartment at 15F, No. 234, Renai Road using mobile phone +886-2-8765-4321 on March 4, 2024, reporting feeling increasingly unwell. Emergency same-day appointment arranged. On presentation, random blood glucose measured at 425 mg/dL (critical value). Direct admission from clinic to hospital ward recommended and accepted by patient. Mr. Smith mentioned living in Taipei since 2010 (14 years), working as English teacher at international school near his Renai Road residence. Email john.smith@hospital.org used for all medical correspondence and appointment reminders since establishing care at TVGH in 2015.

【History of Present Illness - Detailed】
Patient John Smith (MRN-2024-004, DOB 05/12/1963, age 61, contact +886-2-8765-4321) reports progressively worsening symptoms over the past 3 months. Initially noticed increased thirst and urination frequency approximately 12 weeks ago (early December 2023) but attributed symptoms to hot weather and increased coffee consumption. However, symptoms persisted and intensified. Patient reports waking up 5-6 times per night to urinate (nocturia), drinking 4-5 liters of water daily, and experiencing persistent dry mouth despite fluid intake. Unintentional weight loss noted - patient weighed 88 kg in November 2023, now weighs 80 kg (8 kg loss in 3 months). Patient also reports blurred vision for past month, particularly difficulty reading small print and seeing street signs while walking near his apartment on Renai Road (仁愛路). Increased fatigue - previously walked 5 kilometers daily, now feels exhausted after 1 kilometer. Some mild lower extremity numbness and tingling noted for past 2 weeks (possible peripheral neuropathy). No chest pain, no shortness of breath, no abdominal pain, no fever, no recent infections. Patient admits to poor medication compliance over past 6 months due to "feeling fine" and not refilling prescriptions regularly. Last endocrinology follow-up visit was September 2023 (6 months ago) when HbA1c was 7.8%. Patient acknowledges he stopped checking blood glucose at home and stopped taking Metformin regularly since October 2023.

【Past Medical History】
1. Type 2 Diabetes Mellitus: 
   - Diagnosed August 2015 (patient was 52 years old) during routine health check at TVGH
   - Initial HbA1c at diagnosis: 9.2%, fasting glucose 256 mg/dL
   - Initially controlled with Metformin monotherapy
   - 2018: Added Glimepiride due to poor control
   - 2020: Added Sitagliptin (DPP-4 inhibitor) to regimen
   - Baseline HbA1c range 7.5-8.5% with medications (suboptimal control)
   - No prior history of DKA (diabetic ketoacidosis) or HHS (hyperosmolar hyperglycemic state)
   - Patient has not attended diabetes education classes despite multiple recommendations

2. Hypertension:
   - Diagnosed 2012 (age 49) in USA before moving to Taiwan
   - Currently taking Amlodipine 5mg daily
   - Blood pressure generally well-controlled 120-135/75-85 mmHg

3. Hyperlipidemia:
   - Diagnosed 2015 concurrently with diabetes
   - Taking Atorvastatin 20mg nightly
   - LDL goal <100 mg/dL, typically ranges 110-130 mg/dL (suboptimal)

4. Obesity:
   - BMI 28.9 kg/m² (height 175 cm, weight 88 kg prior to recent weight loss)
   - Current BMI 26.1 kg/m² (weight 80 kg) - still overweight category

5. Gastroesophageal Reflux Disease (GERD):
   - Diagnosed 2017, controlled with Omeprazole 20mg daily

6. Benign Prostatic Hyperplasia (BPH):
   - Diagnosed 2020 (age 57), mild symptoms, managed with Tamsulosin 0.4mg nightly

7. Surgical History:
   - Appendectomy 1985 (age 22) in New York, USA - uncomplicated
   - Right knee arthroscopic surgery 2008 (age 45) for meniscal tear - good recovery
   - No other surgeries

8. Family History:
   - Father: John Smith Sr., died age 68 (year 2005) from complications of diabetes and renal failure
   - Mother: Margaret Smith, age 87, living in USA, history of hypertension and osteoporosis
   - One brother: age 59, healthy
   - One sister: age 56, has hypothyroidism
   - Strong family history of diabetes (father, paternal grandfather, paternal aunt)

【Social History】
Patient John Michael Smith (born May 12, 1963 in Manhattan, New York City, USA, currently 61 years old) is a divorced American expatriate who has lived in Taiwan since July 2010 (14 years residence). Works full-time as senior English teacher and coordinator at Taipei American School. Patient resides alone in a 15th floor apartment at No. 234, Renai Road, Section 3, Da'an District, Taipei (仁愛路三段234號15樓). This has been his address since August 2010 when he first moved to Taiwan. The apartment is a 2-bedroom unit with elevator access. Contact phone number +886-2-8765-4321 is both his mobile and home line. Email john.smith@hospital.org used for work and medical communications.

Marital Status: Divorced (2009), no current partner
Children: Two adult children living in USA - daughter Emma (age 33) and son David (age 30), maintains regular contact via video calls weekly
Education: Bachelor of Arts in English Literature (Columbia University, 1985), Master of Education (NYU, 1995)
Occupation: English teacher/coordinator, working at same school since 2010, typical workday 8 AM to 4 PM
Exercise: Previously walked 5 km daily around Da'an Forest Park (大安森林公園) near his apartment on Renai Road, recently decreased due to fatigue
Diet: Admits to poor dietary habits - frequent consumption of takeout food, bubble tea, night market snacks, irregular meal times, large portions
Tobacco: Never smoker
Alcohol: Social drinker - approximately 2-3 beers per week, occasional wine with dinner
Illicit Drugs: Denies
Sleep: 6-7 hours per night, recently disrupted by nocturia (waking 5-6 times to urinate)
Travel: Annual trips to USA to visit children, last trip December 2023 (returned to Taiwan January 2024)
Support System: Several close friends in Taipei expatriate community, no family in Taiwan

【Medication List on Admission】
Patient John Smith (MRN-2024-004, DOB 05/12/1963, insurance INS-US-789456123) reported medications - NOTE: Poor compliance admitted
1. Metformin 1000mg PO BID (prescribed but not taking regularly for past 5-6 months)
2. Glimepiride 4mg PO daily before breakfast (prescribed but not taking regularly)
3. Sitagliptin 100mg PO daily (prescribed but not taking regularly)
4. Amlodipine 5mg PO daily (taking intermittently)
5. Atorvastatin 20mg PO nightly (taking most days)
6. Omeprazole 20mg PO daily (taking most days)
7. Tamsulosin 0.4mg PO at bedtime (taking regularly for BPH symptoms)
8. Aspirin 100mg PO daily (taking intermittently)

Allergies: NO KNOWN DRUG ALLERGIES (NKDA) - confirmed with patient John Smith on March 5, 2024

【Physical Examination on Admission】
Examination Date/Time: March 5, 2024, at 3:00 PM
Location: Endocrinology Ward 6B, Room 605, Bed A, TVGH
Examiner: Dr. Michael Chen (陳明哲醫師) and Dr. Linda Wang (王麗娜醫師, Resident)

Vital Signs:
- Blood Pressure: 148/88 mmHg (elevated, above goal)
- Heart Rate: 92 bpm (mildly tachycardic)
- Respiratory Rate: 18 breaths/minute
- Temperature: 36.7°C (oral)
- Oxygen Saturation: 98% on room air
- Height: 175 cm (5 feet 9 inches)
- Weight: 80 kg (176 pounds) - patient reports 8 kg weight loss from baseline 88 kg
- Body Mass Index: 26.1 kg/m² (overweight, was 28.9 kg/m²)

General Appearance: Mr. John Smith (age 61, born May 12, 1963) is a well-developed, well-nourished Caucasian male appearing stated age. Alert and oriented to person, place, time, and situation. Cooperative with examination. Appears mildly dehydrated with dry mucous membranes. No acute distress.

Skin: Warm and dry. No jaundice, no cyanosis. Multiple seborrheic keratoses on trunk (age-related changes). No diabetic dermopathy, no necrobiosis lipoidica. Skin turgor slightly decreased (consistent with mild dehydration).

Head/HEENT:
- Head: Normocephalic, atraumatic
- Eyes: Pupils equal, round, reactive to light and accommodation (PERRL). Extraocular movements intact (EOMI). Conjunctivae slightly pale. Sclera anicteric. Fundi examination deferred (ophthalmology consult ordered for diabetic retinopathy screening)
- Ears: Tympanic membranes intact bilaterally, normal landmarks
- Nose: Nasal mucosa dry, no discharge
- Throat: Oropharynx without erythema or exudate, mucous membranes dry, poor dentition noted (multiple dental caries - patient reports has not seen dentist in 2 years)

Neck: Supple, no jugular venous distension (JVD), no lymphadenopathy, no thyromegaly, carotid pulses 2+ bilaterally without bruits

Cardiovascular: Regular rate and rhythm (RRR), normal S1 and S2, no murmurs, rubs, or gallops. Point of maximal impulse (PMI) non-displaced. Peripheral pulses: radial 2+ bilateral, dorsalis pedis 1+ bilateral (diminished), posterior tibial 1+ bilateral (diminished) - consistent with possible peripheral arterial disease

Respiratory: Chest expansion symmetric, lungs clear to auscultation bilaterally, no wheezes, rales, or rhonchi, no increased work of breathing

Abdomen: Soft, obese, non-tender, non-distended. Bowel sounds present and normal in all four quadrants. No hepatosplenomegaly on palpation. No masses, no hernias. No costovertebral angle (CVA) tenderness.

Extremities:
- No clubbing, cyanosis, or edema in bilateral upper extremities
- Bilateral lower extremities: No edema, no varicosities
- Feet examination: Bilateral feet warm, pulses diminished as noted above
- Monofilament test: ABNORMAL - decreased sensation to 10g monofilament on plantar surface bilateral great toes and metatarsal heads (evidence of peripheral neuropathy)
- Vibration sense: Decreased at bilateral great toes
- Ankle reflexes: Diminished bilaterally (consistent with neuropathy)
- No foot ulcers, no calluses, no deformities noted

Neurological:
- Mental Status: Alert and oriented x4 (person, place, time, situation)
- Cranial Nerves: II-XII grossly intact
- Motor: Strength 5/5 in bilateral upper and lower extremities
- Sensory: As noted above - decreased sensation in feet (peripheral neuropathy)
- Coordination: Finger-to-nose intact, heel-to-shin intact
- Gait: Steady, no ataxia

【Laboratory Results - Admission March 5, 2024】
Blood Collection Time: 3:30 PM

Complete Blood Count (CBC):
- White Blood Cell Count: 8,500/μL (normal range 4,000-11,000)
- Hemoglobin: 13.2 g/dL (normal male 13.5-17.5, slightly low - possible anemia of chronic disease)
- Hematocrit: 39% (slightly low)
- Mean Corpuscular Volume (MCV): 88 fL (normocytic)
- Platelet Count: 245,000/μL (normal)

Comprehensive Metabolic Panel (CMP):
- Sodium: 148 mEq/L (HIGH - hypernatremia, normal 135-145) - consistent with dehydration
- Potassium: 3.8 mEq/L (normal)
- Chloride: 105 mEq/L (normal)
- Bicarbonate: 19 mEq/L (LOW - metabolic acidosis, normal 22-28) - concerning for DKA
- Blood Urea Nitrogen (BUN): 32 mg/dL (HIGH - prerenal azotemia, normal 7-20)
- Creatinine: 1.3 mg/dL (elevated from baseline 0.9 mg/dL - acute kidney injury from dehydration)
- eGFR: 58 mL/min/1.73m² (Stage 3A CKD - decreased from baseline eGFR 75)
- Glucose: 425 mg/dL (CRITICAL HIGH - normal 70-100 fasting)
- Calcium: 9.5 mg/dL (normal)
- Total Protein: 7.8 g/dL (normal)
- Albumin: 4.0 g/dL (normal)
- AST: 32 U/L (normal)
- ALT: 45 U/L (mildly elevated - possible fatty liver)
- Alkaline Phosphatase: 88 U/L (normal)
- Total Bilirubin: 0.8 mg/dL (normal)

Diabetes-Specific Tests:
- Hemoglobin A1c (HbA1c): 12.5% (VERY HIGH - poor long-term glucose control, goal <7%)
- Fructosamine: 425 μmol/L (elevated, normal <285 - reflects 2-3 week average glucose)
- Serum Ketones: 1.8 mmol/L (elevated, normal <0.6) - mild ketonemia
- Beta-hydroxybutyrate: 2.2 mmol/L (elevated, normal <0.4) - mild ketosis but not DKA range

Lipid Panel (fasting sample):
- Total Cholesterol: 245 mg/dL (HIGH, goal <200)
- LDL Cholesterol: 155 mg/dL (HIGH, goal <100 for diabetic patient)
- HDL Cholesterol: 38 mg/dL (LOW, goal >40 for males)
- Triglycerides: 260 mg/dL (HIGH, goal <150)
- Non-HDL Cholesterol: 207 mg/dL (HIGH, goal <130)

Renal Function:
- Urine Microalbumin: 185 mg/g creatinine (HIGH - microalbuminuria, normal <30, indicates early diabetic nephropathy)
- Spot Urine Protein/Creatinine Ratio: 0.4 (elevated, suggests proteinuria)

Thyroid Function Tests:
- TSH: 2.8 mIU/L (normal 0.4-4.0)
- Free T4: 1.1 ng/dL (normal)

Urinalysis:
- Appearance: Yellow, clear
- Specific Gravity: 1.030 (concentrated - dehydration)
- pH: 5.5 (acidic)
- Glucose: 4+ (>1000 mg/dL) - POSITIVE (glycosuria)
- Ketones: 2+ (moderate) - POSITIVE (ketonuria)
- Protein: 1+ (trace proteinuria)
- Blood: Negative
- Leukocyte Esterase: Negative
- Nitrites: Negative
- Microscopy: No WBCs, no bacteria (no UTI)

Arterial Blood Gas (ABG) - to rule out DKA:
- pH: 7.32 (mildly acidotic, normal 7.35-7.45)
- pCO2: 32 mmHg (low - respiratory compensation)
- pO2: 95 mmHg (normal)
- HCO3: 18 mEq/L (low - metabolic acidosis)
- Anion Gap: 18 (elevated, normal 8-12) - consistent with mild ketoacidosis
INTERPRETATION: Mild metabolic acidosis with respiratory compensation, mild ketonemia - does not meet criteria for DKA but hyperglycemic emergency requiring admission

【Imaging and Diagnostic Studies】
1. Electrocardiogram (ECG) - March 5, 2024, 4:00 PM:
   Normal sinus rhythm at 90 bpm, normal axis, normal intervals (PR 160 ms, QRS 90 ms, QT 400 ms), no ST-segment changes, no Q waves - no acute ischemia

2. Chest X-ray (Portable AP) - March 5, 2024, 5:00 PM:
   Heart size normal, lungs clear, no infiltrates, no effusions, no pneumothorax, no pulmonary edema - no acute cardiopulmonary disease

【Assessment & Plan - Admission March 5, 2024】
Patient: John Michael Smith, 61-year-old American male (born May 12, 1963), MRN-2024-004, residing at 15F, No. 234, Renai Road, Taipei, contact +886-2-8765-4321, email john.smith@hospital.org, insurance INS-US-789456123

PROBLEM LIST:
#1 Uncontrolled Type 2 Diabetes Mellitus with severe hyperglycemia (glucose 425 mg/dL, HbA1c 12.5%)
#2 Mild ketosis with metabolic acidosis (not frank DKA but concerning)
#3 Acute kidney injury secondary to dehydration (Cr 1.3, eGFR 58, hypernatremia)
#4 Diabetic complications: peripheral neuropathy, possible early nephropathy (microalbuminuria), possible retinopathy (needs screening)
#5 Hypertension, suboptimally controlled
#6 Dyslipidemia, poorly controlled
#7 Medication non-compliance (patient admits stopped taking diabetes medications 5-6 months ago)

PLAN:
- Admit to Endocrinology Ward 6B, Room 605, Bed A
- IV fluids: Normal saline 150 mL/hr to correct dehydration and hypernatremia
- Start insulin therapy: Basal-bolus regimen with Insulin Glargine (long-acting) 20 units at bedtime + Insulin Lispro (rapid-acting) per sliding scale before meals
- Frequent blood glucose monitoring: q4-6 hours initially
- Diabetes education consultation: arrange diabetes nurse educator to meet with patient John Smith (contact 02-8765-4321) to discuss medication compliance, diet, exercise
- Nutrition consultation: diabetic diet 1800 kcal/day, low sodium
- Resume home medications: Amlodipine, Atorvastatin, Omeprazole, Tamsulosin
- Hold Metformin temporarily (acute kidney injury)
- Ophthalmology consult for dilated fundus exam (screen for diabetic retinopathy)
- Podiatry consult for foot care education (patient has peripheral neuropathy)
- Social work consult to address medication compliance barriers and support system
- Discharge planning: transition to outpatient insulin if needed, close endocrinology follow-up

Patient John Smith (MRN-2024-004, age 61, DOB 05/12/1963, address Renai Road) counseled extensively regarding importance of medication compliance and risks of uncontrolled diabetes. Patient expressed understanding and commitment to improving diabetes management. Will require intensive education and close follow-up.

【Progress Notes - Hospital Day 1 through Discharge】
(Abbreviated for brevity - patient stayed 5 days, glucose gradually controlled, transitioned to insulin therapy, extensive education provided, discharged home March 10, 2024, with close outpatient follow-up arranged)

DISCHARGE SUMMARY:
Discharge Date: March 10, 2024
Patient: John Michael Smith, Age 61 (born May 12, 1963), MRN-2024-004
Address: 15F, No. 234, Renai Road, Section 3, Taipei | Phone: +886-2-8765-4321 | Email: john.smith@hospital.org | Insurance: INS-US-789456123
Discharge Medications: Insulin Glargine, Insulin Lispro, Metformin (resumed), Amlodipine, Atorvastatin, Aspirin, Omeprazole, Tamsulosin
Follow-up: Endocrinology Clinic March 17, 2024, Ophthalmology March 20, 2024

【Attending Physician】Dr. Michael Chen (陳明哲醫師), MD, PhD, Endocrinology, TVGH | Date: March 10, 2024"""),
        
        ("MRN-2024-005", "Mary Johnson", "1988-09-30", 36, "Female", "+886-4-7654-3210", "mary.j@email.com", "Apt 8B, Lane 156, Zhongxiao E Rd, Taipei", "INS-UK-456123789",
         "Ms. Mary Johnson (female, born September 30, 1988, age 36) visited Chang Gung Hospital on Feb 20, 2024. Contact details: mobile +886-4-7654-3210, email mary.j@email.com. Lives in Apt 8B, Lane 156, Zhongxiao East Road, Taipei. Insurance: INS-UK-456123789. Patient Mary Johnson (born 1988/09/30) can be reached at the above phone or apartment 8B in Lane 156."),
        
        # Japanese names with embedded PHI
        ("MRN-2024-006", "田中太郎 (Tanaka Taro)", "1970-02-18", 54, "男性", "03-9876-5432", "tanaka.t@hospital.jp", "東京都新宿区西新宿2-8-1", "保険-JP-123456789",
         "患者の田中太郎様(生年月日:1970年2月18日、54歳男性)が2024年3月12日に台北日本人クリニックを受診。連絡先:03-9876-5432、メール:tanaka.t@hospital.jp。住所:東京都新宿区西新宿2-8-1。保険番号:保険-JP-123456789。Mr. Tanaka Taro (born Feb 18, 1970) contacted at 03-9876-5432 regarding his condition. Patient 田中 mentioned his address in 西新宿 during intake."),
        
        ("MRN-2024-007", "佐藤花子 (Sato Hanako)", "1985-12-25", 38, "女性", "06-8765-4321", "sato.h@email.jp", "大阪府大阪市北区梅田3-1-3", "保険-JP-987654321",
         "佐藤花子さん(1985年12月25日生まれ、38歳女性)が馬偕醫院婦産科を2024年2月28日に訪問。電話06-8765-4321、Email: sato.h@email.jp。居住地:大阪府大阪市北区梅田3-1-3。保険:保険-JP-987654321。Ms. Sato Hanako (DOB: 12/25/1985) can be contacted via phone 06-8765-4321 or at her residence in 梅田区."),
        
        # Korean names with embedded PHI
        ("MRN-2024-008", "김민준 (Kim Min-jun)", "1995-04-07", 29, "남성", "010-1234-5678", "kim.mj@hospital.kr", "서울특별시 강남구 테헤란로 152", "건보-KR-890123456",
         "환자 김민준님 (생년월일: 1995년 4월 7일, 29세 남성)이 2024년 3월 5일 台北馬偕醫院에 내원하셨습니다. 연락처: 010-1234-5678, 이메일: kim.mj@hospital.kr. 주소: 서울특별시 강남구 테헤란로 152. 건강보험: 건보-KR-890123456. Patient Kim Min-jun (born April 7, 1995) contacted via 010-1234-5678. Mentioned living in 강남구 Gangnam during consultation."),
        
        ("MRN-2024-009", "박서연 (Park Seo-yeon)", "1982-08-14", 42, "여성", "010-9876-5432", "park.sy@email.kr", "부산광역시 해운대구 우동 1408", "건보-KR-567890123",
         "Ms. 박서연 (1982년 8월 14일생, 42세 여성) visited 台中榮總 on Jan 25, 2024. Phone: 010-9876-5432, Email: park.sy@email.kr. Address: 부산광역시 해운대구 우동 1408. Insurance: 건보-KR-567890123. Patient Park Seo-yeon (born Aug 14, 1982) can be reached at mobile 010-9876-5432 or email park.sy@email.kr. Lives in 해운대구 Haeundae."),
        
        # Thai names with embedded PHI
        ("MRN-2024-010", "สมชาย วงศ์ไทย (Somchai Wongthai)", "1968-06-20", 56, "ชาย", "+66-2-123-4567", "somchai.w@hospital.th", "123 ถนนสุขุมวิท แขวงคลองเตย กรุงเทพฯ", "NHSO-TH-1234567890",
         "คนไข้ สมชาย วงศ์ไทย (เกิด 20 มิถุนายน 1968, อายุ 56 ปี เพศชาย) มาโรงพยาบาล台北榮總 วันที่ 15 มกราคม 2024. โทร: +66-2-123-4567, อีเมล: somchai.w@hospital.th. ที่อยู่: 123 ถนนสุขุมวิท แขวงคลองเตย กรุงเทพฯ. บัตรประชาชน: NHSO-TH-1234567890. Mr. Somchai Wongthai (born June 20, 1968) contacted at +66-2-123-4567. Patient mentioned living on Sukhumvit Road (ถนนสุขุมวิท) in Bangkok."),
        
        # Mixed examples with narrative PHI
        ("MRN-2024-011", "李秀英 (Lee Soo-young)", "1991-01-15", 33, "Female", "02-5555-6666", "lee.sy@example.com", "新北市板橋區文化路一段188號", "NHI-D789123456",
         "Patient 李秀英/Lee Soo-young (born 1991/01/15, 33-year-old female) admitted to 亞東醫院 on March 1, 2024. Contact: 02-5555-6666 or lee.sy@example.com. Lives in 新北市板橋區文化路一段188號. Insurance: NHI-D789123456. Ms. Lee (DOB: Jan 15, 1991) mentioned celebrating her 33rd birthday recently during consultation."),
        
        ("MRN-2024-012", "วิชัย เจริญสุข / Vichai Chen", "1975-10-03", 49, "男/Male", "+66-2-987-6543", "vichai.c@email.com", "台北市松山區南京東路五段123號8F", "NHI-E123987456",
         "Mr. วิชัย เจริญสุข (Vichai Chen), born October 3, 1975 (49歲), Thai-Taiwanese patient visited 台大醫院 on Feb 14, 2024. โทร: +66-2-987-6543, Email: vichai.c@email.com. ที่อยู่: 台北市松山區南京東路五段123號8F. 保險: NHI-E123987456. Patient Vichai (born 03/10/1975) can be contacted at phone +66-2-987-6543 or at his apartment 8F on Nanjing E Road (南京東路)."),
        
        # High-risk cases (Age > 90) with detailed narratives - EXTENDED TO >1024 TOKENS
        ("MRN-2024-013", "張老太太", "1928-03-22", 96, "女", "02-1111-2222", "chang.family@email.tw", "台北市士林區中正路456號", "NHI-F456789012",
         """【老年醫學科病歷 Geriatric Medicine Record】⚠️ HIGHLY IDENTIFIABLE - AGE 96
Patient: 張老太太 (Mrs. Chang) | MRN: MRN-2024-013 | DOB: 1928年3月22日 (March 22, 1928) | Age: 96歲 (96 years old) ⚠️ | Gender: 女性 (Female)
Family Contact: 女兒張小芳 (Daughter Chang Xiao-Fang) Phone 02-1111-2222 | Email: chang.family@email.tw
Address: 台北市士林區中正路456號 (No.456, Zhongzheng Rd., Shilin Dist., Taipei City)
Insurance: NHI-F456789012

【Patient Background - 高度可識別風險】
⚠️ CRITICAL PRIVACY NOTE: Patient 張老太太 (Mrs. Chang) was born on March 22, 1928 (民國17年3月22日), making her 96 years old in 2024. This places her in the HIGHLY IDENTIFIABLE age category (>90 years). Combined with her residential address in Shilin District (士林區中正路456號), her age alone significantly increases re-identification risk. Under HIPAA Safe Harbor method, individuals over 89 years must have their ages redacted or reported as ">89 years old." This patient has been residing at the same address on Zhongzheng Road (中正路) for over 60 years since 1960 (民國49年) when she was 32 years old.

【Admission Information 入院資訊】
Admission Date: January 5, 2024, 10:00 AM via Emergency Department
Hospital: 台北榮民總醫院老年醫學科 (Taipei Veterans General Hospital, Department of Geriatric Medicine)
Ward: 老年醫學病房3C (Geriatric Ward 3C) | Room/Bed: Room 305, Bed A
Attending Physician: 王建民主任醫師 (Dr. Wang Jian-Min), Chief of Geriatrics
Admission Diagnosis: Acute functional decline, Failure to thrive, Age-related frailty, Multiple comorbidities

【Chief Complaint & History】
Mrs. 張老太太 (born March 22, 1928, now 96 years old ⚠️) was brought to the Emergency Department by her daughter 張小芳 (age 68, born 1956) on the morning of January 5, 2024. Daughter called ambulance from their shared residence at 台北市士林區中正路456號 (phone 02-1111-2222) around 8:30 AM after finding her mother unable to get out of bed, confused, and refusing to eat breakfast. Patient lives with daughter 張小芳 at No.456 Zhongzheng Road, Shilin District - they have lived at this address together since 1960 when patient張老太太 was 32 years old and recently married, now for 64 years. The house is a traditional two-story building near Shilin Night Market (士林夜市). Email chang.family@email.tw is managed by daughter for medical communications.

Patient 張老太太 (DOB 03/22/1928, age 96) was her usual baseline functional status until approximately 1 week ago (late December 2023). Daughter reports mother was previously able to ambulate with walker independently around the house, feed herself, maintain continence, and carry on conversations. However, over the past week, daughter noticed progressive decline: decreased oral intake, increased confusion (calling daughter by wrong names, thinking it's 1950s), two falls at home (December 30 and January 2), increased lethargy, and urinary incontinence (new onset). Daughter張小芳 states "my 96-year-old mother (born 1928, in the year of Dragon 龍年) has always been sharp mentally and physically active for her age until this week - something is very wrong." No witnessed fever, but daughter felt mother's forehead was warm. No cough, no chest pain, no vomiting, but decreased appetite and fluid intake for 5 days.

【Past Medical History - Extensive due to Age 96】
Patient張老太太 (MRN-2024-013, born March 22, 1928, currently 96 years old, living at 士林區中正路456號 for 64 years) has the following medical conditions:

1. Hypertension: Diagnosed 40+ years ago (approximately 1980s when patient was in her 50s), currently controlled
2. Chronic Kidney Disease Stage 3b: Baseline creatinine 1.5-1.8 mg/dL, eGFR 30-35 mL/min
3. Osteoporosis: Severe, compression fractures of T12 and L1 vertebrae in 2018 (age 90), kyphotic posture
4. Osteoarthritis: Bilateral knees and hands, chronic pain managed conservatively
5. Dementia - Alzheimer's Type: Diagnosed 2020 (age 92), mild to moderate stage, MMSE score 18/30 at last check (October 2023)
6. Atrial Fibrillation: Diagnosed 2015 (age 87), on anticoagulation
7. History of Falls: Multiple falls over past 2 years, high fall risk, uses walker
8. Urinary Incontinence: Stress and urge incontinence, uses pads
9. Constipation: Chronic, requires bowel regimen
10. Hearing Loss: Bilateral, wears hearing aids
11. Cataracts: Bilateral, s/p cataract surgery both eyes (2012 age 84)
12. Anemia of Chronic Disease: Baseline Hb 10-11 g/dL

Surgical History:
- Cesarean Section x2 (1955 and 1958 when patient was 27 and 30 years old) - delivered daughter 張小芳 via C-section in 1956 at age 28
- Cholecystectomy (gallbladder removal) 1995 (age 67)
- Cataract surgeries: Right eye 2012, Left eye 2012 (both at age 84)

Family History:
- Patient張老太太 born in 1928, is one of the oldest individuals in her family lineage
- Parents: Both deceased (father died 1960s age ~60s, mother died 1970s age ~70s)
- Siblings: Had 3 siblings, all deceased (patient is sole survivor of her generation in family)
- Husband: Mr. Chang (張先生) deceased 2005 at age 79 (married 1955-2005, 50 years of marriage)
- Children: Two daughters - 張小芳 (age 68, born 1956, primary caregiver) and 張小梅 (age 66, born 1958, lives abroad in USA)
- Grandchildren: Five grandchildren, ages 35-45
- Great-grandchildren: Eight great-grandchildren, youngest age 2

【Social History - Patient Age 96, Born 1928】
Mrs. 張老太太 (Chang, born March 22, 1928, age 96) has lived an exceptionally long life spanning nearly a century. Born in 1928 during Japanese colonial rule of Taiwan (日治時期), she experienced Japanese occupation, World War II as a teenager (was 17 years old when WWII ended in 1945), the transition to ROC governance, and witnessed Taiwan's transformation over 96 years. Patient received elementary education during Japanese rule, speaks Japanese, Taiwanese, and some Mandarin. Married Mr. 張先生 in 1955 at age 27, had two daughters. Worked as seamstress for 30+ years until retirement around 1990 (age 62). Widowed in 2005 at age 77 when husband died.

Current Living Situation:
Patient 張老太太 (age 96, DOB 1928-03-22) resides with her daughter 張小芳 (age 68) at the family home located at 台北市士林區中正路456號 (No.456, Zhongzheng Road, Shilin District, Taipei). This has been the family residence since 1960 (64 years) - patient moved here with husband in 1960 when she was 32 years old, raised two daughters in this house, and has remained here ever since. The address 中正路456號 is well-known to neighbors as the Chang family residence. Two-story traditional house, patient's bedroom on first floor to avoid stairs. Daughter張小芳 is primary caregiver, retired nurse. Foreign domestic helper (外籍看護) from Indonesia assists with ADLs (activities of daily living) during daytime hours Monday-Friday. Contact number 02-1111-2222 is home landline, has been the Chang family phone number for decades.

Functional Status Prior to Admission:
- Ambulation: Required walker for household mobility, high fall risk, unable to go outdoors independently
- ADLs (Activities of Daily Living): Required assistance with bathing, dressing, toileting; able to self-feed with setup
- IADLs (Instrumental ADLs): Fully dependent on daughter for cooking, cleaning, medication management, finances
- Cognitive Status: Mild-moderate dementia (MMSE 18/30), forgetful, occasionally confused about time/place, but usually recognizes family
- Social: Enjoys sitting in front yard watching street activity, visits from grandchildren, watching Taiwanese opera (歌仔戲) on TV

Diet: Soft diet due to poor dentition (few remaining teeth), prefers traditional Taiwanese food
Tobacco: Never smoker
Alcohol: Never used alcohol
Sleep: 8-10 hours per night, frequent daytime naps

【Medications Prior to Admission】
Home medications for patient 張老太太 (MRN-2024-013, born 1928-03-22, age 96):
1. Amlodipine 5mg PO daily (hypertension) - taking for 20+ years
2. Furosemide 20mg PO daily (heart failure prevention, CKD)
3. Warfarin 2mg PO daily (atrial fibrillation) - INR monitored monthly, last INR 2.3 (December 2023)
4. Donepezil 5mg PO at bedtime (Alzheimer's dementia) - started 2020 age 92
5. Calcium 600mg + Vitamin D 400 IU PO BID (osteoporosis)
6. Multivitamin PO daily
7. Sennosides 17.2mg PO at bedtime PRN constipation
8. Acetaminophen 500mg PO q6h PRN pain (for arthritis)

Allergies: Penicillin - causes rash (reported allergy from 1960s, patient was ~35 years old when discovered)

【Physical Examination - Emergency Department January 5, 2024】
Exam Time: 11:00 AM | Location: TVGH Emergency Department
Patient: 張老太太, Female, 96 years old (born March 22, 1928), appears frail and older than stated age

Vital Signs:
- Blood Pressure: 158/82 mmHg (elevated from baseline 130-140/70-80)
- Heart Rate: 108 bpm (tachycardia, irregularly irregular - atrial fibrillation)
- Respiratory Rate: 24 breaths/min (tachypnea)
- Temperature: 38.3°C (101°F) - FEVER
- Oxygen Saturation: 91% on room air (hypoxemia) → improved to 95% on 2L nasal cannula
- Weight: 42 kg (92 pounds) - patient张老太太 baseline weight 45 kg, lost 3 kg over past month
- Height: 145 cm (4 feet 9 inches) - shortened from younger height due to kyphosis and vertebral compression fractures

General Appearance: 
Mrs. 張老太太 (age 96, DOB 1928-03-22, from 士林區中正路456號) appears as a very elderly, frail, cachectic Asian female appearing older than stated age of 96 years. Lethargic but arousable to voice. Oriented to self only (knows her name), not oriented to place, time, or situation. Daughter at bedside providing history. Patient appears ill and dehydrated.

Skin: Thin, fragile skin with multiple ecchymoses (bruising) on bilateral arms and legs (likely from recent falls and friable skin). Senile purpura present. Skin turgor poor (tenting). Multiple seborrheic keratoses. No pressure ulcers noted (good skin care by daughter).

HEENT:
- Head: Normocephalic, sparse white hair, atraumatic
- Eyes: Sunken eyes (dehydration), pupils equal and reactive, bilateral cataracts (s/p surgery), conjunctivae pale (anemia)
- Ears: Bilateral hearing aids in place but not turned on
- Nose: Dry nasal mucosa
- Mouth: Edentulous (no teeth), dry mucous membranes (dehydration), no oral lesions

Neck: Supple, no jugular venous distension, no lymphadenopathy, no thyromegaly

Cardiovascular: Irregularly irregular rhythm (atrial fibrillation), tachycardic rate ~108 bpm, normal S1/S2, II/VI systolic murmur at apex (chronic, known murmur), no rubs or gallops. Peripheral pulses: radial 2+ bilateral, dorsalis pedis 1+ bilateral, posterior tibial 1+ bilateral.

Respiratory: Tachypneic at 24 breaths/min. Decreased breath sounds at bilateral bases. Crackles (rales) in right lower lung field. Dullness to percussion right base. No wheezes.

Abdomen: Soft, non-distended, hypoactive bowel sounds, mild diffuse tenderness without guarding or rebound, no masses, no organomegaly

Genitourinary: Foley catheter placed in ED draining concentrated, cloudy, foul-smelling urine (~50 mL initial output) - concerning for urinary tract infection (UTI)

Musculoskeletal: Severe kyphosis (hunchback deformity from vertebral compression fractures), bilateral knee crepitus, limited range of motion all joints due to osteoarthritis and deconditioning, muscle wasting (sarcopenia)

Extremities: No edema, multiple bruises as noted, skin fragile

Neurological:
- Mental Status: Lethargic, Glasgow Coma Scale 14/15 (E4V4M6), oriented to self only, not oriented to place/time/situation (baseline dementia MMSE 18/30, but more confused than usual per daughter)
- Cranial Nerves: Difficult to assess due to cognitive impairment and lethargy, pupils reactive
- Motor: Generalized weakness, muscle atrophy, unable to stand without assistance, moving all extremities
- Sensory: Decreased sensation in bilateral feet (peripheral neuropathy)
- Reflexes: Diminished throughout
- Gait: Deferred (patient unable to stand)

【Laboratory Results - January 5, 2024】
Complete Blood Count:
- WBC: 15,800/μL (HIGH - leukocytosis, concerning for infection)
- Hemoglobin: 9.2 g/dL (LOW - anemia, decreased from baseline 10-11 g/dL)
- Hematocrit: 28% (LOW)
- Platelets: 178,000/μL (low-normal)

Comprehensive Metabolic Panel:
- Sodium: 152 mEq/L (HIGH - hypernatremia from dehydration)
- Potassium: 4.8 mEq/L (high-normal)
- Chloride: 108 mEq/L (high-normal)
- Bicarbonate: 20 mEq/L (LOW - metabolic acidosis)
- BUN: 58 mg/dL (HIGH - prerenal azotemia)
- Creatinine: 2.4 mg/dL (HIGH - acute on chronic kidney injury, baseline 1.5-1.8)
- eGFR: 18 mL/min/1.73m² (Stage 4 CKD acutely, baseline 30-35)
- Glucose: 165 mg/dL (mildly elevated, stress hyperglycemia)
- Calcium: 8.5 mg/dL (low-normal)

Liver Function Tests: Within normal limits for age

Coagulation Studies:
- INR: 3.2 (elevated from goal 2.0-3.0 - supratherapeutic warfarin, explains bruising)
- PT: 32 seconds (prolonged)
- PTT: 35 seconds (normal)

Urinalysis:
- Appearance: Cloudy, foul-smelling
- Color: Dark amber
- Specific Gravity: 1.025 (concentrated)
- pH: 6.0
- Protein: 2+ (proteinuria)
- Glucose: Negative
- Ketones: Trace
- Blood: 2+ (hematuria)
- Leukocyte Esterase: 3+ POSITIVE (infection)
- Nitrites: POSITIVE (bacteria)
- WBCs: >100/hpf (pyuria - infection)
- Bacteria: Many bacteria seen
INTERPRETATION: Urinary Tract Infection (UTI)

Urine Culture: Sent, pending (later grew Escherichia coli >100,000 CFU/mL, sensitive to Ceftriaxone)

Blood Cultures: Two sets sent (later negative x2)

Chest X-ray (Portable): Right lower lobe infiltrate/consolidation consistent with pneumonia. Small right pleural effusion. Mild cardiomegaly. Osteopenia/severe osteoporosis noted.

ECG: Atrial fibrillation with rapid ventricular response, rate 108 bpm, no acute ST changes

【Assessment & Plan - Admission】
Patient: 張老太太 (Mrs. Chang), Female, 96 years old ⚠️ (DOB: March 22, 1928), MRN: MRN-2024-013
Address: 台北市士林區中正路456號 (lived at this address for 64 years since 1960)
Contact: Daughter 張小芳 at 02-1111-2222, email chang.family@email.tw
Insurance: NHI-F456789012

PROBLEM LIST:
#1 Acute Functional Decline / Failure to Thrive - in 96-year-old elderly patient
#2 Community-Acquired Pneumonia (right lower lobe) - CURB-65 score 4 (high risk)
#3 Urinary Tract Infection (UTI) with pyuria - E. coli
#4 Acute Kidney Injury (AKI) on Chronic Kidney Disease Stage 3b - Cr 2.4 (baseline 1.5-1.8)
#5 Dehydration with Hypernatremia (Na 152) - poor oral intake x 5 days
#6 Supratherapeutic INR 3.2 on Warfarin (atrial fibrillation) - bleeding risk
#7 Age-related Frailty - patient 96 years old, multiple comorbidities
#8 Alzheimer's Dementia - baseline MMSE 18/30, now with delirium superimposed on dementia
#9 Anemia - Hb 9.2 (decreased from baseline 10-11) - likely anemia of inflammation from infections
#10 Multiple comorbidities as listed in PMH (HTN, AFib, osteoporosis, etc.)

PLAN:
- Admit to Geriatric Ward 3C, Room 305, Bed A
- Goals of Care Discussion: Daughter 張小芳 (phone 02-1111-2222) is healthcare proxy, discussed prognosis given patient's age 96 and acute illnesses, daughter requests full medical treatment at this time but no CPR if cardiac arrest (DNR/Full Treatment)
- IV Fluids: Normal saline 75 mL/hr (gentle rate due to age 96 and CKD) - correct dehydration and hypernatremia gradually
- Antibiotics: Ceftriaxone 1g IV daily (renally dosed for CKD) for pneumonia + UTI coverage
- Hold Warfarin temporarily (INR 3.2, supratherapeutic) - restart when INR <3.0
- Oxygen: Continue 2L nasal cannula, goal SpO2 >92%
- Nutrition: NPO initially, advance to soft diet as tolerated
- PT/OT consult: Assess function, rehabilitation potential given age 96 and frailty
- Geriatric Syndrome Management: Fall prevention, delirium prevention, skin care, bowel/bladder management
- Family Meeting: Schedule with daughter 張小芳 to discuss prognosis and discharge planning for 96-year-old mother
- Monitor: Daily labs (renal function, electrolytes), vital signs q4h, strict I/O

PROGNOSIS: Guarded given patient's advanced age (96 years old, born 1928), multiple acute processes (pneumonia, UTI, AKI), and baseline frailty. Will require skilled nursing facility level care likely upon discharge. Daughter張小芳 aware of serious nature of illness in her 96-year-old mother.

【Subsequent Hospital Course - Days 1-14】
Patient 張老太太 (96 years old, born March 22, 1928) had prolonged hospitalization complicated by delirium, slow recovery from pneumonia given age, and significant functional decline. Required 14-day hospitalization. Antibiotics completed 10-day course. Kidney function improved to near baseline (Cr 1.9). Delirium gradually improved. Physical therapy involved but patient remained very deconditioned. Daughter張小芳 (02-1111-2222) visited daily.

Discharge Date: January 19, 2024
Discharge Disposition: Transferred to Skilled Nursing Facility for rehabilitation and skilled care
Discharge Diagnosis: 
(1) Community-acquired pneumonia, resolved
(2) Urinary tract infection, treated
(3) Acute kidney injury on CKD, improved
(4) Failure to thrive in 96-year-old patient
(5) Advanced age with frailty

Patient 張老太太 (age 96, DOB 1928-03-22, previously living at 士林區中正路456號 for 64 years) no longer able to return home safely. Arranged transfer to skilled nursing facility near daughter's home in Shilin District. Daughter張小芳 contact 02-1111-2222 for all future medical decisions.

【主治醫師】Dr. 王建民 (Wang Jian-Min), Geriatric Medicine, TVGH | Date: January 19, 2024

⚠️ PRIVACY NOTE: This patient's age (96 years) and long-term residence at 士林區中正路456號 (64 years at same address) make her HIGHLY RE-IDENTIFIABLE. Proper de-identification MUST redact/modify age to ">89" and consider address generalization."""),
        
        ("MRN-2024-014", "山田一郎 (Yamada Ichiro)", "1932-11-30", 92, "男性", "03-2222-3333", "yamada.family@jp.com", "台北市大同區迪化街一段234號", "保険-JP-234567890",
         "患者 山田一郎様 (Yamada Ichiro), 生年月日 1932年11月30日、現在92歳の男性。2024年2月1日に台大醫院神経内科に入院。連絡先: 03-2222-3333、メール: yamada.family@jp.com。住所: 台北市大同區迪化街一段234號。保険: 保険-JP-234567890。⚠️ HIGH-RISK: Age 92. Mr. Yamada (born Nov 30, 1932) has lived in 迪化街 Dihua Street since 1950s. Contact family at 03-2222-3333."),
        
        # Rare disease cases with sensitive information
        ("MRN-2024-015", "劉小華", "1998-05-18", 26, "女", "04-6666-7777", "liu.xh@rare.org.tw", "台中市北區三民路三段99號", "NHI-G789012345",
         "Patient 劉小華, female, DOB 1998-05-18 (26歲), diagnosed with Fabry Disease (法布瑞氏症) - RARE DISEASE at 台中榮總罕見疾病中心 on Jan 10, 2024. ⚠️ RARE DISEASE - highly identifiable! Contact: 04-6666-7777, liu.xh@rare.org.tw. Address: 台中市北區三民路三段99號. Insurance: NHI-G789012345. Ms. Liu (born May 18, 1998) mentioned her rare diagnosis and regular enzyme replacement therapy. Lives in 三民路 Sanmin Road."),
        
        ("MRN-2024-016", "เด็กหญิง ดาวใจ / Daowjai", "2010-07-12", 14, "หญิง/Female", "+66-2-777-8888", "daowjai.parent@email.th", "台北市內湖區成功路四段188號", "NHI-H012345678",
         "Pediatric patient เด็กหญิง ดาวใจ (Daowjai), born July 12, 2010 (14歲 female), diagnosed with Pompe Disease (龐貝氏症) - RARE PEDIATRIC DISEASE at 馬偕醫院小兒罕病科 on Feb 15, 2024. ⚠️ RARE DISEASE + MINOR! Parent contact: +66-2-777-8888, daowjai.parent@email.th. Address: 台北市內湖區成功路四段188號. Insurance: NHI-H012345678. Young patient Daowjai (DOB: 07/12/2010, age 14) receives weekly ERT. Thai family living in 內湖區 Neihu District."),
    ]
    
    # Add more random patients to reach 50
    additional_names = [
        ("陳雅芳", "台灣"),
        ("Jennifer Lee", "美國"),
        ("佐々木誠 (Sasaki Makoto)", "日本"),
        ("최지우 (Choi Ji-woo)", "韓國"),
        ("สุภาพ แสงทอง", "泰國"),
    ]
    
    for i in range(17, 51):
        name_data = random.choice(additional_names)
        age = random.randint(20, 95)
        birth_year = 2024 - age
        birth_month = random.randint(1,12)
        birth_day = random.randint(1,28)
        dob = f"{birth_year}-{birth_month:02d}-{birth_day:02d}"
        mrn = f"MRN-2024-{i:03d}"
        gender = random.choice(["男", "女", "Male", "Female"])
        phone = f"0{random.randint(2,9)}-{random.randint(1000,9999)}-{random.randint(1000,9999)}"
        email = f"patient{i}@email.com"
        district = random.choice(['大安', '信義', '中山', '內湖', '松山'])
        address = f"台北市{district}區測試路{random.randint(1,999)}號"
        insurance = f"NHI-{chr(65+random.randint(0,25))}{random.randint(100000000,999999999)}"
        
        # Generate narrative with embedded PHI
        narrative = f"Patient {name_data[0]} (born {dob}, age {age}, {gender}) visited hospital on 2024-0{random.randint(1,3)}-{random.randint(10,28)}. Contact: {phone} or {email}. Lives in {address} ({district}區). Insurance: {insurance}. Patient mentioned being {age} years old and living in {district} District during intake. Can be reached at phone {phone}."
        
        patients.append((
            mrn,
            name_data[0],
            dob,
            age,
            gender,
            phone,
            email,
            address,
            insurance,
            narrative
        ))
    
    # Write patient data
    for row_idx, patient in enumerate(patients, 2):
        for col_idx, value in enumerate(patient, 1):
            cell = ws1.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Auto-size columns (10 columns now including narrative)
    for col in range(1, 10):
        ws1.column_dimensions[chr(64+col)].width = 20
    # Make narrative column wider
    ws1.column_dimensions['J'].width = 80
    
    ws1.row_dimensions[1].height = 45
    
    # Sheet 2: Clinical Visits (臨床就診記錄)
    ws2 = wb.create_sheet("Clinical Visits")
    
    visit_headers = [
        "Visit ID", "Patient ID", "Visit Date\n就診日期", 
        "Hospital\n醫院", "Ward\n病房", "Bed\n床號",
        "Doctor\n主治醫師", "Diagnosis\n診斷", "Treatment\n治療",
        "Clinical Notes\n臨床記錄"  # NEW: Narrative notes with embedded PHI
    ]
    
    for col, header in enumerate(visit_headers, 1):
        cell = ws2.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    visits = [
        ("V-001", "MRN-2024-001", "2024-01-15", "台大醫院 / NTUH", "心臟內科病房 5A", "Bed-512", 
         "Dr. 李文華", "Hypertension (高血壓)", "Amlodipine 5mg",
         "Patient 陳建國 (MRN-2024-001, age 69, DOB 1955-03-15) admitted to Ward 5A, Bed 512 on Jan 15, 2024. Dr. 李文華 noted BP 165/95. Contact: 02-2234-5678. Patient lives at 台北市大安區信義路四段123號5樓. Started on Amlodipine 5mg. Follow-up scheduled. Call 02-2234-5678 to confirm appointment."),
        
        ("V-002", "MRN-2024-006", "2024-01-20", "台北榮總 / TVGH", "神經內科 7B", "Bed-703",
         "Dr. 鈴木一郎 / Dr. Suzuki", "Parkinson's Disease (帕金森氏症)", "Levodopa",
         "田中太郎様 (Tanaka Taro, MRN-2024-006, born 1970-02-18, 54歳) admitted to Neurology Ward 7B, Bed 703 on 2024/01/20. Dr. 鈴木一郎診察。Contact: 03-9876-5432, tanaka.t@hospital.jp. Address: 東京都新宿区西新宿2-8-1. Started Levodopa treatment for Parkinson's."),
        
        ("V-003", "MRN-2024-013", "2024-02-01", "三軍總醫院 / TSGH", "老年醫學科 3C", "Bed-305",
         "Dr. 王建民", "Age-related frailty (老年衰弱症) - Age 96", "Rehabilitation",
         "⚠️ HIGH-RISK AGE 96: 張老太太 (Mrs. Chang, MRN-2024-013, born 1928-03-22, currently 96 years old) admitted to Geriatrics Ward 3C, Bed 305. Family contact: 02-1111-2222. Patient lives at 台北市士林區中正路456號. Dr. 王建民 notes severe frailty in 96-year-old patient. Rehab plan initiated. Contact daughter at 02-1111-2222."),
        
        ("V-004", "MRN-2024-015", "2024-02-10", "台中榮總 / TCVGH", "罕見疾病中心 2F", "Bed-201",
         "Dr. 陳美玲", "Fabry Disease (法布瑞氏症) - Rare", "Enzyme replacement therapy",
         "⚠️ RARE DISEASE: 劉小華 (Ms. Liu, MRN-2024-015, born 1998-05-18, age 26) admitted to Rare Disease Center, 2F, Bed 201 on Feb 10. Diagnosed with Fabry Disease (法布瑞氏症) - extremely rare condition. Contact: 04-6666-7777, liu.xh@rare.org.tw. Lives at 台中市北區三民路三段99號. Dr. 陳美玲 started ERT. Patient Liu (26歲) with rare Fabry Disease."),
        
        ("V-005", "MRN-2024-008", "2024-02-15", "서울대병원 / Seoul National Hospital", "내과 4층", "침대-401",
         "Dr. 김철수", "Type 2 Diabetes (제2형 당뇨병)", "Metformin 500mg",
         "환자 김민준 (Kim Min-jun, MRN-2024-008, 생년월일 1995-04-07, 29세) 내과 4층 침대 401호에 2024년 2월 15일 입원. 연락처: 010-1234-5678, kim.mj@hospital.kr. 주소: 서울특별시 강남구 테헤란로 152. Dr. 김철수 진료. Metformin 500mg 처방. Patient Kim (born Apr 7, 1995) lives in 강남구 Gangnam."),
        
        ("V-006", "MRN-2024-010", "2024-03-01", "โรงพยาบาลจุฬาลงกรณ์ / Chulalongkorn Hospital", "หอผู้ป่วย 6A", "เตียง-615",
         "Dr. วิชัย สุขใจ", "Dengue Fever (ไข้เลือดออก)", "Supportive care",
         "คนไข้ สมชาย วงศ์ไทย (Somchai Wongthai, MRN-2024-010, เกิด 1968-06-20, อายุ 56 ปี) เข้าพักหอผู้ป่วย 6A เตียง 615 วันที่ 1 มีนาคม 2024. โทร: +66-2-123-4567, somchai.w@hospital.th. ที่อยู่: 123 ถนนสุขุมวิท กรุงเทพฯ. Dr. วิชัย สุขใจ วินิจฉัย Dengue Fever. Mr. Somchai (born June 20, 1968) from Sukhumvit Road."),
        
        ("V-007", "MRN-2024-014", "2024-03-10", "台北慈濟醫院 / Tzu Chi Hospital", "老人科 8B", "Bed-802",
         "Dr. 林志明", "Dementia (失智症) - Age 92", "Donepezil + Care plan",
         "⚠️ HIGH-RISK AGE 92: 山田一郎様 (Yamada Ichiro, MRN-2024-014, 生年月日 1932-11-30, 92歳) 老人科 8B病棟 Bed 802に入院。連絡: 03-2222-3333, yamada.family@jp.com. 住所: 台北市大同區迪化街一段234號. Dr. 林志明 diagnosed dementia in 92-year-old patient. Donepezil started. Contact family at 03-2222-3333."),
        
        ("V-008", "MRN-2024-016", "2024-03-20", "馬偕醫院 / Mackay Hospital", "小兒罕病科 2A", "Bed-210",
         "Dr. 張小芬", "Pompe Disease (龐貝氏症) - Rare pediatric", "ERT weekly",
         "⚠️ RARE DISEASE + MINOR: เด็กหญิง ดาวใจ (Daowjai, MRN-2024-016, born 2010-07-12, age 14) admitted to Pediatric Rare Disease Ward 2A, Bed 210. Pompe Disease (龐貝氏症) - rare pediatric condition. Parent contact: +66-2-777-8888, daowjai.parent@email.th. Address: 台北市內湖區成功路四段188號. Dr. 張小芬 managing weekly ERT for 14-year-old Daowjai with rare Pompe Disease."),
    ]
    
    for row_idx, visit in enumerate(visits, 2):
        for col_idx, value in enumerate(visit, 1):
            cell = ws2.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    for col in range(1, 10):
        ws2.column_dimensions[chr(64+col)].width = 25
    # Make clinical notes column wider
    ws2.column_dimensions['J'].width = 80
    
    # Sheet 3: Lab Results (檢驗報告)
    ws3 = wb.create_sheet("Lab Results")
    
    lab_headers = [
        "Lab ID", "Patient ID", "Test Date", "Test Name\n檢查項目",
        "Result\n結果", "Unit", "Reference Range\n參考值", "Status"
    ]
    
    for col, header in enumerate(lab_headers, 1):
        cell = ws3.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    labs = [
        ("LAB-001", "MRN-2024-001", "2024-01-15", "Blood Pressure (血壓)", "165/95", "mmHg", "120/80", "High"),
        ("LAB-002", "MRN-2024-006", "2024-01-20", "Alpha-synuclein (α-突觸核蛋白)", "Elevated", "ng/mL", "Normal", "Abnormal"),
        ("LAB-003", "MRN-2024-013", "2024-02-01", "Frailty Index (衰弱指數)", "0.45", "score", "<0.25", "Severe"),
        ("LAB-004", "MRN-2024-015", "2024-02-10", "Alpha-galactosidase A", "0.8", "nmol/hr/mg", "1.5-7.5", "Low"),
        ("LAB-005", "MRN-2024-008", "2024-02-15", "HbA1c (당화혈색소)", "8.2", "%", "<5.7", "High"),
        ("LAB-006", "MRN-2024-016", "2024-03-20", "GAA enzyme (acid α-glucosidase)", "<1", "nmol/hr/mg", "3-15", "Critical"),
    ]
    
    for row_idx, lab in enumerate(labs, 2):
        for col_idx, value in enumerate(lab, 1):
            cell = ws3.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    for col in range(1, 9):
        ws3.column_dimensions[chr(64+col)].width = 22
    
    # Save file
    filename = "data/test/test_medical_records_multilang.xlsx"
    wb.save(filename)
    print(f"✅ Generated: {filename}")
    print(f"   - 50 patients with PHI in 5 languages")
    print(f"   - Multiple sheets with clinical data")
    print(f"   - High-risk cases: Age >90, Rare diseases")
    return filename


if __name__ == "__main__":
    generate_test_excel_multilanguage()
