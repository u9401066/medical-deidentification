#!/usr/bin/env python3
"""
Complex PHI Test Data Generator - Mixed Column Challenge
生成複雜個資混雜測試數據 - 提高去識別化難度

設計理念:
1. 減少明確可刪除的個資欄位（如純姓名欄、純電話欄）
2. 將個資混入業務欄位中（如診斷描述、醫囑、護理記錄）
3. 增加隱性個資（年齡推算、地理位置線索、罕見疾病）
4. 跨欄位個資關聯（同一個資在不同欄位以不同形式出現）
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate_complex_phi_test():
    """生成個資樣態複雜的測試數據"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Complex PHI Cases"

    # 表頭設計 - 混合欄位，不能直接刪除
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        "Case ID\n案例編號",
        "Clinical Summary\n臨床摘要\n(含姓名/年齡/診斷)",
        "Contact & Demographics\n聯絡與人口學\n(混合資訊)",
        "Medical History\n病史\n(含時間線/地點)",
        "Treatment Notes\n治療記錄\n(含醫師/日期/處置)",
        "Social Context\n社會情境\n(職業/居住/家庭)",
        "Risk Indicators\n風險指標\n(年齡/罕病/遺傳)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 設置列寬
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 40

    # 案例數據 - 各種個資樣態
    cases = [
        # Case 1: 高齡患者 (Age > 90) - 多重時間線索
        (
            "CASE-001",
            "Patient 陳老先生 (Mr. Chen), 94-year-old male presenting with hip fracture. Born in 1930 during Japanese colonial period. Chief complaint: fell at home yesterday (Nov 21, 2024) and unable to stand.",
            "Emergency contact: Son 陳大衛 at 02-2758-9999. Lives alone before fall. Previous address verification: 台北市信義區 resident since 1960s. Medicare ID: A123456789 (issued 1995 when patient turned 65).",
            "PMH: Diagnosed with AFib in 2015 (age 85), HTN since 1995 (age 65), had appendectomy in 1945 (age 15 during WWII). Patient recalls being treated by Dr. 王at台大醫院 for TB in 1950 (age 20).",
            "Admitted Nov 21, 2024 to Ortho Ward 5A Bed 512. Dr. 林建國 performed hemiarthroplasty Nov 22. Post-op day 1: stable. Patient mentioned he graduated from 台北高校 in 1948 (age 18). Discharge planned Nov 28.",
            "Retired teacher, taught at 建國中學 for 40 years (1955-1995, ages 25-65). Widower since 2010 (wife passed at age 82). Has 3 children: eldest son born 1958 (when patient was 28), daughter born 1962, youngest son born 1965.",
            "⚠️ HIGH-RISK AGE: 94 years old (DOB 1930). Patient born in Year of Horse 馬年. Mentioned living through 1945 bombing of Taipei as 15-year-old. Re-identification risk: VERY HIGH."
        ),

        # Case 2: 罕見疾病 + 年輕患者
        (
            "CASE-002",
            "Ms. 李雅婷 (Lee Ya-Ting), 28-year-old female with Huntington's Disease (舞蹈症). Patient DNA test showed CAG repeat 43 (abnormal). Mother died age 52 from same disease (2018). Presents with chorea and behavioral changes.",
            "Cell: 0912-345-678. Email: yalee1996@gmail.com. Emergency: Father 李先生 0933-888-999. Insurance: Catastrophic Illness Card #HD-2023-0056 (issued Jan 2023 age 27). Genetic counseling consent signed.",
            "Family Hx: Mother had HD (diagnosed 2005 age 39, died 2018 age 52). Maternal grandmother also affected (died 1995 age 68). Patient tested age 25 (2021) after mother's death. Brother age 30 tested negative. Lives 新竹市東區.",
            "Baseline UHDRS motor score: 35 (moderate). Started Tetrabenazine 25mg BID on Mar 15, 2024 by Dr. 張神經科. Brain MRI shows caudate atrophy. Patient seen at 台大神經部 HD clinic every 3 months since 2023.",
            "Social worker: Patient lost job as software engineer (worked at 新竹科學園區 2019-2023) due to symptoms. Applied for disability in Feb 2024 age 28. Lives with father. Engaged but wedding postponed after diagnosis.",
            "⚠️ RARE DISEASE: HD prevalence ~1/10,000. Age 28 + HD + Hsinchu = highly identifiable. Genetic: CAG 43 repeats. Family: 3 generations affected (grandmother-mother-patient). Predicted onset age 35-40, actual 25."
        ),

        # Case 3: 小兒罕病 + 監護人資訊
        (
            "CASE-003",
            "Pediatric patient 小明 (Xiao-Ming), 8-year-old boy with Duchenne Muscular Dystrophy (DMD). Born May 15, 2016. Started walking late (18 months), diagnosis confirmed age 5 (2021) via genetic testing (DMD gene deletion exons 45-52).",
            "Parents: 王爸爸 (father) 0988-123-456, 王媽媽 (mother) 0955-789-012. Home: 台中市西屯區 near 逢甲大學. School: 西屯國小 2nd grade. Mother quit job to care full-time. Father works at 台中工業區.",
            "Dx: Motor delay noted age 2 (2018), frequent falls age 4. CPK elevated 15,000 U/L (normal <200). Genetic test 2021 showed deletion. Wheelchair-bound since age 7 (2023). Cardiac echo: early cardiomyopathy.",
            "Treatment by Dr. 陳at中榮小兒神經科: Steroid (Prednisolone 0.75mg/kg daily) started 2022 age 6. PT/OT 3x/week. Ataluren trial considered but not eligible (nonsense mutation required, patient has deletion). Annual cardiac monitoring.",
            "Only child. Mother age 35 (born 1989), father age 38 (born 1986) - both carriers. Maternal uncle had DMD, died age 19 (2005). Parents considering IVF/PGD for second child. Family receives 罕病基金會 support.",
            "⚠️ RARE PEDIATRIC: DMD ~1/5,000 boys. Age 8 + DMD + Taichung + specific deletion = identifiable. Family: X-linked, maternal uncle affected. Carrier mother quit job 2023. Wheelchair since age 7. Progressive disease."
        ),

        # Case 4: 職業性疾病 + 地理位置
        (
            "CASE-004",
            "Mr. 黃文雄 (Huang Wen-Hsiung), 58yo male, retired coal miner with pneumoconiosis (塵肺症). Worked 麥寮六輕 1985-2020 (ages 23-58, 35 years). CXR shows progressive massive fibrosis. DOB: Feb 8, 1966.",
            "Addr: 雲林縣麥寮鄉橋頭村 (near plant). Tel: 05-6921-XXX. Occupational disease certified 2020. Labor insurance: 00-123456. Spouse: 黃太太 (age 56). Son works same plant.",
            "Exposure: Chemical operator 1985-2005 (20 yrs), then maintenance 2005-2020 (15 yrs). No respirator use before 2000. Smoking: 40 pack-years (quit 2020 age 58). Asbestos exposure confirmed in job records.",
            "Pulmonary function: FEV1 45% predicted (severe obstruction). Chest CT 2024: bilateral upper lobe masses concerning for lung CA vs PMF. Biopsy scheduled. Seen by Dr. 李胸腔科 at 台大雲林分院 since 2020.",
            "Disability: Unable to work since 2020 age 58. Filed for occupational disease compensation (approved 2021). Receives NT$35,000/month disability pension. Lives in company dormitory (since 1985, 39 years same address).",
            "Risk: Pneumoconiosis + 40 pack-year smoking = high lung CA risk. Location: 麥寮六輕 worker cluster (multiple similar cases). Age at retirement 58 (early due to disease). Same address 39 years. Son same occupation."
        ),

        # Case 5: 精神科 + 法律案件
        (
            "CASE-005",
            "Patient 林某 (Mr. Lin), 35yo male with schizophrenia, admitted involuntarily after suicide attempt Nov 10, 2024. Jumped from 台北101 observation deck (85F), survived with bilateral calcaneal fractures. Prior: 5 psych admissions 2015-2024.",
            "Court order: Mandatory treatment per Mental Health Act. Public defender: 張律師 02-2999-8888. ID: A12345678X. Contact: Mother 林媽媽 (age 62) 0922-111-222. No fixed address (homeless intermittently since 2020).",
            "Psychiatric: First episode psychosis age 22 (2011), diagnosed schizophrenia 2012. Multiple hospitalizations: 2015(age 32), 2017, 2019, 2021, 2023, current. Poor medication compliance. Substance use: amphetamine (last use Nov 2024).",
            "Current: Transferred from Ortho to Psych ward Nov 15 after surgery. On Risperdal 4mg BID, Depakote 500mg BID. Denies SI/HI now. Psychiatrist: Dr. 王精神科. Case manager: 陳社工 coordinates with 台北市衛生局.",
            "Social: Dropped out 台大資工系 2013 (age 25) due to illness. Unemployed since. Estranged from father. Mother retired nurse, sole support. Was living 萬華區 homeless shelter before incident. Has 重大傷病卡.",
            "Legal: Police report #2024-11-10-0123. Attempted suicide at public landmark (高度可識別事件). Media reported 'man jumped from 101' Nov 10. Forensic eval ordered. Criminal charges pending vs mental health diversion."
        ),

        # Case 6: 產科 + 高齡產婦
        (
            "CASE-006",
            "Gravida 1 Para 0, 王小姐 (Ms. Wang), 46-year-old primigravida at 28 weeks gestation via IVF. AMA (advanced maternal age). Amnio: Trisomy 21 (Down syndrome). Couple considering TOP vs continue. EDC: Feb 20, 2025.",
            "IVF at 送子鳥診所 Taipei, 3rd cycle successful (prior 2 failed). Husband age 52 (張先生). Tel: 0916-888-999. Addr: 天母 high-rise. Both professionals (王=lawyer, 張=doctor). Insurance: private + NHI.",
            "Obstetric: LMP May 15, 2024. IVF-ET June 10, 2024. NIPT at 10 weeks: high risk T21. Amnio at 16 weeks confirmed (karyotype 47,XX,+21). Fetal echo: AVSD (heart defect). Growth: 10th percentile.",
            "Maternal: IDDM (gestational diabetes on insulin). BP 145/95 (mild preeclampsia labs negative). Seen by Dr. 林婦產科 at 台安醫院 high-risk OB clinic biweekly. Pediatric genetics consult: counseled on DS prognosis.",
            "Decision-making: Couple struggled with decision for 12 weeks. Religious (Christian, attend 信友堂). Family pressure to terminate. Finally decided to continue pregnancy Nov 2024. Planning for NICU delivery, cardiac surgery.",
            "Risk factors: Age 46 (T21 risk 1/30 at this age vs 1/700 at age 30). IVF pregnancy (3rd cycle). Primigravida at advanced age. IDDM + preeclampsia. Fetus: T21 + AVSD. Identifiable: 天母 + age 46 + IVF + T21 diagnosis."
        ),

        # Case 7: 器官移植 + 捐贈者
        (
            "CASE-007",
            "Liver transplant recipient 陳先生 (Mr. Chen), 55yo male with HBV cirrhosis, MELD 28, listed May 2024. Donor: deceased 28yo motorcycle accident victim (DOA Nov 15, 2024 at 彰化基督教醫院). OPO case #2024-11-15-001.",
            "Recipient: ID T123456789, Blood O+, Tel 04-2233-4455, Lives 彰化市. Donor: Brain death declared Nov 15, 21:00. Family (parents age 55/53) consented to multi-organ donation. Liver, kidneys, heart, corneas procured.",
            "Surgical: Orthotopic liver transplant performed Nov 16, 2024, 02:00-08:00 (6 hrs) by Dr. 李移植外科 team at TCVGH. Cold ischemic time: 6.5 hrs. Warm ischemic time: 45 min. Blood loss: 3,500 mL, transfused 12U PRBC.",
            "Post-op: ICU course complicated by primary non-function, required re-transplant Nov 20 (2nd donor). Second graft functioning. Immunosuppression: Tacrolimus, MMF, Prednisone. Rejection episode POD#7 treated with steroid pulse.",
            "Donor info (sensitive): 28yo male from 南投縣, engineering student at 中興大學, accident at 台14線 mountain road Nov 15, 06:00. Next of kin: parents 陳爸爸/陳媽媽. Donor liver split: recipient + pediatric recipient.",
            "Ethical/Legal: Organ donation consent #12345. UNOS/TORSC allocation. Recipient waited 6 months (expedited due to high MELD). Required 2 donors (1st liver failed). Donor family grief counseling provided. Media interest (young donor)."
        ),

        # Case 8: 愛滋病 + 結核病
        (
            "CASE-008",
            "Anonymous patient, MSM (men who have sex with men), age 32, newly diagnosed HIV+ (Nov 2024) and active pulmonary TB. CD4 85 cells/μL, VL 250,000 copies/mL. Presented with cough, fever, night sweats x 1 month.",
            "Contact tracing: 3 sexual partners past 6 months (notified by 疾管署). Lives 西門町 area. Works at 酒吧 (bar). Denies IDU. Last HIV test 2 years ago was negative. Tel: 0933-XXX-XXX (confidential). Case manager: CDC Taipei.",
            "TB: CXR bilateral infiltrates, sputum AFB+, GeneXpert MTB+ RIF-sensitive. Started RIPE therapy Nov 18. Respiratory isolation. Contact investigation: 5 close contacts (roommates, coworkers) - all TST to be done.",
            "HIV: Diagnosed via rapid test at 昆明街 anonymous testing site Nov 5, confirmed WB+ Nov 10. Genotype: CRF07_BC (common in Taiwan MSM). No OIs except TB. ART planned after TB meds stabilized (will start Biktarvy).",
            "Social: Single, disclosed to mother only (father doesn't know). Sex work history (occasional). Prior STIs: Syphilis (treated 2022), gonorrhea (2023). PrEP never used. Depression diagnosed 2021, on Prozac. Substance: poppers, occasionally ketamine.",
            "Public Health: Reportable diseases (HIV + TB). Partner notification challenging (anonymous encounters). Stigma concerns. TB infectiousness: isolated until culture negative (2-4 weeks). HIV: U=U after virologic suppression. High-risk population."
        ),

        # Case 9: 醫療糾紛 + 不良事件
        (
            "CASE-009",
            "Medical error case: 劉女士 (Ms. Liu), 68yo underwent laparoscopic cholecystectomy Nov 1, 2024. Bile duct injury during surgery (CBD transected). Required conversion to open + hepaticojejunostomy. Prolonged hospitalization (45 days).",
            "Patient: ID B234567890, lives 桃園市中壢區. Family: Husband 劉先生 (age 70) filed complaint Nov 30. Lawyer: 林律師事務所 03-4567-8888. Hospital: 林口長庚. Surgeon: Dr. 王 (general surgery, 15 years experience).",
            "Timeline: Surgery Nov 1, 08:00-11:30 (3.5 hrs, planned 1.5 hrs). Intraop: CBD mistaken for cystic duct, transected. Recognized immediately, hepatobiliary team called (Dr. 陳). Repair: Roux-en-Y hepaticojejunostomy.",
            "Complications: Post-op bile leak, sepsis, ICU 12 days. ERCP stent placed Nov 10. Re-operation Nov 15 for abscess drainage. Total 4 surgeries. Discharged Dec 15, 2024. Estimated additional cost: NT$2,000,000. Medical board review scheduled.",
            "Investigation: Incident report filed Nov 1 (mandatory). Root cause analysis: Lack of intraop cholangiogram, surgeon fatigue (8th case that day), inadequate supervision (fellow performing, attending scrubbed in late). Patient records subpoenaed.",
            "Legal: Medical negligence lawsuit filed Dec 2024. Damage claims: NT$5M. Hospital liability insurance activated. Media coverage (reported in 蘋果日報 Dec 5). Surgeon currently on administrative leave. Case pending mediation vs litigation."
        ),

        # Case 10: 新生兒 + 遺傳代謝疾病
        (
            "CASE-010",
            "Newborn female, 李小妹 (Baby Lee), born Nov 20, 2024 at 36+3 weeks via C/S for fetal distress. Birth weight 2,340g. Newborn screen positive for MCAD deficiency (medium-chain acyl-CoA dehydrogenase, 中鏈脂肪酸代謝異常).",
            "Parents: 李爸爸 (age 33, engineer), 李媽媽 (age 31, teacher), both from 宜蘭縣. Consanguinity: NO (parents unrelated). Tel: 03-9XX-XXXX. Baby in NICU at 羅東博愛醫院. Genetics consult requested.",
            "Newborn Screen: Collected DOL#2 (Nov 22), positive C8 (octanoylcarnitine) elevated. Confirmatory: Plasma acylcarnitine profile abnormal, urine organic acids normal. Genetic test pending (ACADM gene sequencing).",
            "Management: Avoid fasting >4 hrs (risk hypoglycemia, metabolic crisis). High carb, low fat diet. Cornstarch supplementation. Emergency protocol card issued. Parents trained on sick-day management. Metabolic dietitian: 陳營養師.",
            "Family: First child for couple. Mother had prior miscarriage 2022 (unknown cause). MCAD is AR (autosomal recessive), both parents carriers (25% recurrence risk each pregnancy). Genetic counseling provided Nov 25. Family planning discussed.",
            "Risk: MCAD can cause sudden death in infancy if undiagnosed (fasting → hypoglycemia → coma). Newborn screen saved life (pre-symptomatic diagnosis). Parents need education. Identifiable: 宜蘭 + MCAD (rare) + birth date Nov 20, 2024."
        ),

        # Case 11: 跨國醫療 + 外籍患者
        (
            "CASE-011",
            "Foreign patient: NGUYEN Van Thanh (阮文清), 45yo Vietnamese male, migrant worker, fall from scaffold at construction site 台中港區 Nov 12, 2024. Traumatic brain injury (SDH), comatose, GCS 6. No family in Taiwan.",
            "Employer: 鴻海建設 (contact: 工地主任 04-2XXX-XXXX). Insurance: Foreign worker insurance + 勞保. Passport: AB1234567 (Vietnam). Taiwan work permit: expires Dec 2025. No local emergency contact. Vietnamese Economic Cultural Office notified.",
            "Injury: Fall from 5th floor (15 meters), landed on concrete. CT head: large right SDH, midline shift 8mm, skull fracture. Emergent craniotomy by Dr. 張神外 Nov 12. ICU since. Prognosis: guarded, likely permanent disability if survives.",
            "Social: In Taiwan since 2020 (4 years). Sends money to family in Vietnam (wife, 2 children ages 12, 9). Lives in dormitory with 20 other workers. Coworkers visited initially. Language barrier (speaks limited Mandarin, Vietnamese interpreter needed).",
            "Ethical dilemmas: Family in Vietnam cannot afford to come (airfare ~NT$30,000). Video call with wife arranged Nov 18. Patient unresponsive. Goals of care discussion via interpreter. Wife says continue treatment. Who pays if worker uninsured portion?",
            "Outcome: Patient arrested Nov 28 (PEA), resuscitated, now brain dead. Family wants body repatriated to Vietnam (cost ~NT$200,000). Employer negotiating. Organ donation discussed but cultural objections. Case highlights migrant worker healthcare gaps."
        ),

        # Case 12: 整形外科 + 身分辨識特徵
        (
            "CASE-012",
            "Cosmetic surgery patient: 張小姐 (Ms. Chang), 26yo female influencer (Instagram: @beautychang, 500K followers), underwent rhinoplasty + double eyelid surgery Nov 1, 2024 at 美麗境界診所 Taipei. Post-op infection and nasal necrosis.",
            "Patient: Lives 信義區豪宅 (luxury apt), drives white Tesla (license plate: ABC-1234 seen in clinic parking). Tel: 0922-XXX-XXX. Paid cash NT$250,000. Medical tourism (flew from HK for surgery). Posted pre-op photos on IG Oct 30.",
            "Procedure: Rhinoplasty (silicone implant L-shape), double eyelid (incisional method). Surgeon: Dr. 李 (Korea-trained). Surgery 3 hours, uneventful. Discharged same day with antibiotics (Cephalexin), pain meds. Follow-up scheduled Nov 7.",
            "Complication: Patient developed fever, nasal pain Nov 5 (POD#4). Returned to clinic Nov 6: nasal tip black (necrosis), purulent discharge. Admitted to hospital. IV antibiotics (Vancomycin + Ceftriaxone). Implant removed emergently Nov 7. Wound debridement.",
            "Outcome: Salvaged nasal tip but permanent deformity. Patient devastated (livelihood dependent on appearance). Threatened lawsuit, posted negative review. Clinic offered refund + revision surgery free. Patient now off social media (deleted IG Nov 20).",
            "Identity: Highly identifiable (public figure + specific procedures + dates + location). Recognizable face (influencer). Vehicle plate. Clinic name. Rhinoplasty necrosis (rare complication ~0.5%). Professional/financial impact. Reputational concerns both sides."
        ),

        # Case 13: 長期照護 + 失智症
        (
            "CASE-013",
            "Mrs. 黃老太太 (Huang), 88yo female with advanced Alzheimer's dementia (CDR 3), admitted from nursing home 安心養護中心 (台北市文山區) for aspiration pneumonia Nov 18, 2024. Lives in NH since 2020 (age 84), before that lived with daughter 黃小芳.",
            "Family: Primary caregiver: Daughter 黃小芳 (age 62, retired teacher), Tel 02-8661-XXXX. Son 黃大明 (age 65, lives USA). Patient widowed 2015 (husband died age 90). Has 4 grandchildren. Decision-maker: daughter (power of attorney since 2021).",
            "Cognitive: MMSE 3/30 (severe). Non-verbal since 2023. ADL: totally dependent (feeding, bathing, toileting). PEG tube placed 2022 age 86 for dysphagia/aspiration risk. Tube feeding: Ensure 1500 cal/day. Contracted (flexion contractures both legs).",
            "Current: Aspiration PNA (RLL infiltrate CXR). Temp 38.9°C, hypoxic (SpO2 88% on RA). Started Unasyn IV. DNR/DNI order (signed by daughter 2021). Comfort-focused care per family wishes. NH will accept back after acute treatment completed.",
            "Goals of care: Daughter states 'Mom would not want prolonged suffering. She always said when it's time, let her go peacefully.' No CPR, no ICU, no intubation. OK with antibiotics for comfort. Hospice evaluation requested. Case manager coordinating.",
            "Psychosocial: Daughter visits NH 3x/week (lives nearby 景美). Feels guilty placing mother in NH but couldn't provide 24/7 care. Son in USA unable to visit (visited last time 2022). NH cost NT$50,000/month. Daughter pays from mother's pension/savings."
        ),

        # Case 14: 職業運動員 + 公眾人物
        (
            "CASE-014",
            "Professional athlete: 林志傑 (Lin Chih-Chieh), 32yo baseball player for 中信兄弟隊 (CTBC Brothers), underwent Tommy John surgery (UCL reconstruction) Nov 5, 2024 at 高雄長庚 by Dr. 陳運醫 (team physician). Public figure (media covered injury).",
            "Identity: Jersey #25, pitcher, ERA 3.25 this season. Injured during game vs 統一獅 Oct 28, 2024 at 台南球場 (broadcast on CPBL TV, video of injury online). MRI Oct 30: complete UCL tear. Contract: NT$8M/year, expires 2025. Agent: 王經紀人.",
            "Surgery: UCL reconstruction using palmaris longus graft (ipsilateral forearm). Procedure 2.5 hrs. Post-op: elbow brace, sling. Physical therapy protocol: 12-18 months to return to pitching. No guarantee of full recovery (success rate ~85% in MLB).",
            "Media: Press release Nov 6 (team announced surgery). Apple Daily headline: '林志傑動刀 球季報銷' (Lin undergoes surgery, season over). Patient unhappy with lack of privacy (reporters called hospital, photos leaked of him in wheelchair).",
            "Career impact: Age 32, recovery 12-18 months → age 33-34 for return. Contract negotiations: team may not renew. Considering retirement if recovery poor. Financial concerns (mortgage on 高雄豪宅 NT$30M). Wife + 2 kids (ages 5, 3).",
            "Privacy violation: Public figure but medical records protected. Hospital investigating leak (photo posted on PTT). Patient considering legal action. CPBL sent letter to hospital re: confidentiality. High-profile case tests medical privacy in sports."
        ),

        # Case 15: 毒品濫用 + 刑事案件
        (
            "CASE-015",
            "Patient 陳某 (Mr. Chen), 41yo male, arrested Nov 10, 2024 by police for drug trafficking (methamphetamine). Brought to ER in custody with chest pain. Urine drug screen: positive amphetamine, methamphetamine. ECG: ST elevation → STEMI from cocaine (also positive).",
            "Legal: Criminal case #2024-訴-1234, Taipei District Court. Police custody at Taipei Detention Center (台北看守所). Guards present during hospitalization (handcuffed to bed). Lawyer: public defender 林律師. Charges: trafficking (5-15 years), possession.",
            "Medical: STEMI (anterior wall), underwent emergent cardiac cath Nov 10. LAD 100% occlusion (thrombus), PTCA + stent. Post-cath: stable. Cardiology: likely drug-induced vasospasm + thrombosis. History: admits using meth 10+ years, IV cocaine past 2 years.",
            "Psychiatric: Depression, PTSD (childhood trauma). Prior suicide attempt 2019. Substance use disorder (severe). No psychiatric treatment (non-compliant). Social: unemployed, estranged from family. Prior incarcerations: 2015 (6 months), 2018 (1 year).",
            "Disposition: Transferred to prison hospital (法務部醫院) Nov 15 after cardiac stabilization. Continued care there. Scheduled for trial Jan 2025. Addiction treatment in prison (methadone program). Prognosis: high recidivism without treatment.",
            "Forensic: Medical records subpoenaed for trial (drug use evidence). Doctor testified at preliminary hearing. Blood samples stored as evidence. Complex case: Patient vs prisoner vs person with addiction/mental illness. Treatment vs punishment debate."
        )
    ]

    # 填充數據
    for row_idx, case_data in enumerate(cases, start=2):
        for col_idx, value in enumerate(case_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # 高風險案例用紅色標註
            if "⚠️" in str(value) or "HIGH-RISK" in str(value):
                cell.font = Font(color="FF0000")

    # 保存文件
    output_dir = "data/test"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "test_complex_phi_cases.xlsx")
    wb.save(output_file)

    print(f"✅ Generated: {output_file}")
    print("   - 15 complex PHI cases (各種個資樣態)")
    print("   - Mixed columns (no direct deletable PHI columns)")
    print("   - High-risk cases: Age >90, rare diseases, public figures")
    print("   - Challenge: PHI embedded in clinical narratives")
    print("\n📊 PHI Types Included:")
    print("   ✓ Age >90 (Case 1)")
    print("   ✓ Rare diseases (Cases 2, 3, 10)")
    print("   ✓ Genetic information (Cases 2, 3, 7, 10)")
    print("   ✓ Mental health (Case 5)")
    print("   ✓ HIV/AIDS (Case 8)")
    print("   ✓ Substance abuse (Cases 8, 15)")
    print("   ✓ Criminal records (Cases 5, 9, 15)")
    print("   ✓ Public figures (Cases 12, 14)")
    print("   ✓ Occupational identifiers (Case 4)")
    print("   ✓ Geographic identifiers (All cases)")
    print("   ✓ Dates (admission, surgery, events)")
    print("   ✓ Family relationships")
    print("   ✓ Contact information (embedded in text)")

if __name__ == "__main__":
    generate_complex_phi_test()
