#!/usr/bin/env python3
"""
PHI Tagged Test Data Generator - For Precision/Recall Calculation
生成帶 PHI 標記的測試數據 - 用於計算檢出率

標記格式: 【PHI:TYPE:ID】content【/PHI】
- TYPE: PHI 類型 (NAME, AGE, DATE, PHONE, ID, LOCATION, etc.)
- ID: 唯一識別碼 (用於追蹤同一個 PHI)
- content: 實際的 PHI 內容

這樣可以：
1. 精確計算真陽性 (TP): 系統檢出的 PHI ∩ 標記的 PHI
2. 計算假陽性 (FP): 系統檢出但未標記的
3. 計算假陰性 (FN): 標記但未檢出的
4. 計算 Precision = TP / (TP + FP)
5. 計算 Recall = TP / (TP + FN)
6. 計算 F1 Score = 2 * (Precision * Recall) / (Precision + Recall)
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def phi_tag(phi_type, content, phi_id=None):
    """
    為 PHI 添加標記
    
    Args:
        phi_type: PHI 類型 (NAME, AGE, DATE, PHONE, etc.)
        content: PHI 內容
        phi_id: 可選的 PHI ID（用於追蹤重複出現的同一個 PHI）
    
    Returns:
        標記後的字符串
    """
    if phi_id:
        return f"【PHI:{phi_type}:{phi_id}】{content}【/PHI】"
    else:
        return f"【PHI:{phi_type}】{content}【/PHI】"

def generate_phi_tagged_test():
    """生成帶 PHI 標記的測試數據"""

    wb = Workbook()
    ws = wb.active
    ws.title = "PHI Tagged Cases"

    # 表頭設計
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        "Case ID",
        "Clinical Summary\n(含標記的 PHI)",
        "Contact Info\n(含標記的聯絡資訊)",
        "Medical History\n(含標記的時間/地點)",
        "Treatment Notes\n(含標記的醫師/日期)",
        "PHI Count\nPHI 數量統計"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 設置列寬
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 30

    # 案例數據 - 帶 PHI 標記
    cases = [
        # Case 1: 高齡患者
        (
            "CASE-001",
            f"Patient {phi_tag('NAME', '陳老先生', 'P001')} (Mr. {phi_tag('NAME', 'Chen', 'P001')}), "
            f"{phi_tag('AGE_OVER_89', '94-year-old', 'A001')} male presenting with hip fracture. "
            f"Born in {phi_tag('DATE', '1930', 'D001')} during Japanese colonial period. "
            f"Chief complaint: fell at home yesterday ({phi_tag('DATE', 'Nov 21, 2024', 'D002')}) and unable to stand.",

            f"Emergency contact: Son {phi_tag('NAME', '陳大衛', 'P002')} at {phi_tag('PHONE', '02-2758-9999', 'T001')}. "
            f"Lives alone before fall. Previous address verification: {phi_tag('LOCATION', '台北市信義區', 'L001')} "
            f"resident since {phi_tag('DATE', '1960s', 'D003')}. Medicare ID: {phi_tag('ID', 'A123456789', 'ID001')} "
            f"(issued {phi_tag('DATE', '1995', 'D004')} when patient turned {phi_tag('AGE', '65', 'A002')}).",

            f"PMH: Diagnosed with AFib in {phi_tag('DATE', '2015', 'D005')} (age {phi_tag('AGE', '85', 'A003')}), "
            f"HTN since {phi_tag('DATE', '1995', 'D006')} (age {phi_tag('AGE', '65', 'A004')}), had appendectomy in "
            f"{phi_tag('DATE', '1945', 'D007')} (age {phi_tag('AGE', '15', 'A005')} during WWII). Patient recalls being "
            f"treated by Dr. {phi_tag('NAME', '王', 'P003')} at {phi_tag('LOCATION', '台大醫院', 'L002')} for TB in "
            f"{phi_tag('DATE', '1950', 'D008')} (age {phi_tag('AGE', '20', 'A006')}).",

            f"Admitted {phi_tag('DATE', 'Nov 21, 2024', 'D009')} to {phi_tag('LOCATION', 'Ortho Ward 5A Bed 512', 'L003')}. "
            f"Dr. {phi_tag('NAME', '林建國', 'P004')} performed hemiarthroplasty {phi_tag('DATE', 'Nov 22', 'D010')}. "
            f"Post-op day 1: stable. Discharge planned {phi_tag('DATE', 'Nov 28', 'D011')}.",

            "PHI Count: NAME=4, AGE_OVER_89=1, AGE=5, DATE=11, PHONE=1, ID=1, LOCATION=3 | Total=26"
        ),

        # Case 2: 罕見疾病
        (
            "CASE-002",
            f"Ms. {phi_tag('NAME', '李雅婷', 'P005')} ({phi_tag('NAME', 'Lee Ya-Ting', 'P005')}), "
            f"{phi_tag('AGE', '28-year-old', 'A007')} female with Huntington's Disease. "
            f"Patient DNA test showed CAG repeat 43 (abnormal). Mother died age {phi_tag('AGE', '52', 'A008')} "
            f"from same disease ({phi_tag('DATE', '2018', 'D012')}). Presents with chorea and behavioral changes.",

            f"Cell: {phi_tag('PHONE', '0912-345-678', 'T002')}. Email: {phi_tag('EMAIL', 'yalee1996@gmail.com', 'E001')}. "
            f"Emergency: Father {phi_tag('NAME', '李先生', 'P006')} {phi_tag('PHONE', '0933-888-999', 'T003')}. "
            f"Insurance: Catastrophic Illness Card #{phi_tag('ID', 'HD-2023-0056', 'ID002')} "
            f"(issued {phi_tag('DATE', 'Jan 2023', 'D013')} age {phi_tag('AGE', '27', 'A009')}). "
            f"Lives {phi_tag('LOCATION', '新竹市東區', 'L004')}.",

            f"Family Hx: Mother had HD (diagnosed {phi_tag('DATE', '2005', 'D014')} age {phi_tag('AGE', '39', 'A010')}, "
            f"died {phi_tag('DATE', '2018', 'D015')} age {phi_tag('AGE', '52', 'A011')}). Maternal grandmother also "
            f"affected (died {phi_tag('DATE', '1995', 'D016')} age {phi_tag('AGE', '68', 'A012')}). "
            f"Patient tested age {phi_tag('AGE', '25', 'A013')} ({phi_tag('DATE', '2021', 'D017')}) after mother's death. "
            f"Brother age {phi_tag('AGE', '30', 'A014')} tested negative.",

            f"Started Tetrabenazine {phi_tag('DATE', 'Mar 15, 2024', 'D018')} by Dr. {phi_tag('NAME', '張神經科', 'P007')}. "
            f"Patient seen at {phi_tag('LOCATION', '台大神經部', 'L005')} HD clinic every 3 months since "
            f"{phi_tag('DATE', '2023', 'D019')}. Lost job at {phi_tag('LOCATION', '新竹科學園區', 'L006')} "
            f"{phi_tag('DATE', '2019-2023', 'D020')}.",

            "PHI Count: NAME=3, AGE=8, DATE=9, PHONE=2, EMAIL=1, ID=1, LOCATION=3 | Total=27"
        ),

        # Case 3: 小兒罕病
        (
            "CASE-003",
            f"Pediatric patient {phi_tag('NAME', '小明', 'P008')} ({phi_tag('NAME', 'Xiao-Ming', 'P008')}), "
            f"{phi_tag('AGE', '8-year-old', 'A015')} boy with Duchenne Muscular Dystrophy (DMD). "
            f"Born {phi_tag('DATE', 'May 15, 2016', 'D021')}. Started walking late ({phi_tag('AGE', '18 months', 'A016')}), "
            f"diagnosis confirmed age {phi_tag('AGE', '5', 'A017')} ({phi_tag('DATE', '2021', 'D022')}) via genetic testing.",

            f"Parents: {phi_tag('NAME', '王爸爸', 'P009')} (father) {phi_tag('PHONE', '0988-123-456', 'T004')}, "
            f"{phi_tag('NAME', '王媽媽', 'P010')} (mother) {phi_tag('PHONE', '0955-789-012', 'T005')}. "
            f"Home: {phi_tag('LOCATION', '台中市西屯區', 'L007')} near {phi_tag('LOCATION', '逢甲大學', 'L008')}. "
            f"School: {phi_tag('LOCATION', '西屯國小', 'L009')} 2nd grade. "
            f"Father works at {phi_tag('LOCATION', '台中工業區', 'L010')}.",

            f"Motor delay noted age {phi_tag('AGE', '2', 'A018')} ({phi_tag('DATE', '2018', 'D023')}), "
            f"frequent falls age {phi_tag('AGE', '4', 'A019')}. Genetic test {phi_tag('DATE', '2021', 'D024')} showed deletion. "
            f"Wheelchair-bound since age {phi_tag('AGE', '7', 'A020')} ({phi_tag('DATE', '2023', 'D025')}).",

            f"Treatment by Dr. {phi_tag('NAME', '陳', 'P011')} at {phi_tag('LOCATION', '中榮小兒神經科', 'L011')}: "
            f"Steroid started {phi_tag('DATE', '2022', 'D026')} age {phi_tag('AGE', '6', 'A021')}. "
            f"Mother age {phi_tag('AGE', '35', 'A022')} (born {phi_tag('DATE', '1989', 'D027')}), "
            f"father age {phi_tag('AGE', '38', 'A023')} (born {phi_tag('DATE', '1986', 'D028')}). "
            f"Maternal uncle had DMD, died age {phi_tag('AGE', '19', 'A024')} ({phi_tag('DATE', '2005', 'D029')}).",

            "PHI Count: NAME=5, AGE=10, DATE=9, PHONE=2, LOCATION=6 | Total=32"
        ),

        # Case 4: 簡化案例 - 電話和身份證
        (
            "CASE-004",
            f"Patient {phi_tag('NAME', '張三', 'P012')}, age {phi_tag('AGE', '45', 'A025')}, "
            f"admitted on {phi_tag('DATE', '2024-11-20', 'D030')} for routine checkup.",

            f"Contact: {phi_tag('PHONE', '02-1234-5678', 'T006')}. "
            f"ID: {phi_tag('ID', 'A123456789', 'ID003')}. "
            f"Address: {phi_tag('LOCATION', '台北市大安區', 'L012')}.",

            f"Previous visit: {phi_tag('DATE', '2023-11-15', 'D031')}. "
            f"Referred by Dr. {phi_tag('NAME', '李四', 'P013')}.",

            f"Follow-up scheduled: {phi_tag('DATE', '2024-12-20', 'D032')}.",

            "PHI Count: NAME=2, AGE=1, DATE=3, PHONE=1, ID=1, LOCATION=1 | Total=9"
        ),

        # Case 5: Email 和網址
        (
            "CASE-005",
            f"Patient {phi_tag('NAME', 'John Smith', 'P014')}, {phi_tag('AGE', '52', 'A026')} years old, "
            f"English speaker, admitted {phi_tag('DATE', 'Nov 23, 2024', 'D033')}.",

            f"Email: {phi_tag('EMAIL', 'john.smith@example.com', 'E002')}. "
            f"Mobile: {phi_tag('PHONE', '+886-912-345-678', 'T007')}. "
            f"Emergency: Wife {phi_tag('NAME', 'Mary Smith', 'P015')} at {phi_tag('PHONE', '+886-933-111-222', 'T008')}.",

            f"Insurance: {phi_tag('ID', 'INS-2024-9876', 'ID004')}. "
            f"Passport: {phi_tag('ID', 'US123456789', 'ID005')}. "
            f"Temporary address: {phi_tag('LOCATION', '台北市中山區民生東路123號', 'L013')}.",

            f"Seen by Dr. {phi_tag('NAME', 'Wang', 'P016')} on {phi_tag('DATE', 'Nov 23', 'D034')}. "
            f"Next appointment: {phi_tag('DATE', 'Dec 15, 2024', 'D035')}.",

            "PHI Count: NAME=3, AGE=1, DATE=3, PHONE=2, EMAIL=1, ID=2, LOCATION=1 | Total=13"
        ),
    ]

    # 寫入數據
    for row_idx, case in enumerate(cases, 2):
        for col_idx, value in enumerate(case, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # PHI Count 列用不同顏色
            if col_idx == 6:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.font = Font(bold=True, size=9)

    # 保存文件
    output_dir = "data/test"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "test_phi_tagged_cases.xlsx")
    wb.save(output_file)

    print(f"✅ Generated: {output_file}")
    print("   - 5 test cases with PHI tags")
    print("   - Tag format: 【PHI:TYPE:ID】content【/PHI】")
    print("\n📊 PHI Statistics:")
    print("   Case 1: 26 PHI instances (high-risk age >89)")
    print("   Case 2: 27 PHI instances (rare disease)")
    print("   Case 3: 32 PHI instances (pediatric)")
    print("   Case 4:  9 PHI instances (simple)")
    print("   Case 5: 13 PHI instances (international)")
    print("   Total: 107 PHI instances")
    print("\n💡 Usage:")
    print("   1. Parse tags with regex: 【PHI:(\\w+):?(\\w*)】([^【]+)【/PHI】")
    print("   2. Compare with system output to calculate:")
    print("      - True Positives (TP): Correctly detected PHI")
    print("      - False Positives (FP): Incorrectly detected")
    print("      - False Negatives (FN): Missed PHI")
    print("   3. Calculate metrics:")
    print("      - Precision = TP / (TP + FP)")
    print("      - Recall = TP / (TP + FN)")
    print("      - F1 Score = 2 * P * R / (P + R)")

    # 同時生成一個解析腳本範例
    parser_script = """#!/usr/bin/env python3
# -*- coding: utf-8 -*-
\"\"\"
PHI Tag Parser - Extract Ground Truth PHI
解析 PHI 標記，提取標準答案
\"\"\"

import re
import pandas as pd
from typing import List, Dict, Tuple

def parse_phi_tags(text: str) -> List[Dict]:
    \"\"\"
    從文本中解析所有 PHI 標記
    
    Returns:
        List of dict with keys: type, id, content, start, end
    \"\"\"
    pattern = r'【PHI:(\\w+):?(\\w*)】([^【]+?)【/PHI】'
    matches = []
    
    for match in re.finditer(pattern, text):
        phi_type = match.group(1)
        phi_id = match.group(2) if match.group(2) else None
        content = match.group(3)
        
        matches.append({
            'type': phi_type,
            'id': phi_id,
            'content': content,
            'start': match.start(),
            'end': match.end(),
            'full_match': match.group(0)
        })
    
    return matches

def calculate_metrics(ground_truth: List[Dict], detected: List[Dict]) -> Dict:
    \"\"\"
    計算檢出率指標
    
    Args:
        ground_truth: 標記的 PHI 列表
        detected: 系統檢出的 PHI 列表
    
    Returns:
        Dict with TP, FP, FN, Precision, Recall, F1
    \"\"\"
    # 轉換為集合方便比較 (使用 content 作為鍵)
    gt_set = {(phi['type'], phi['content']) for phi in ground_truth}
    det_set = {(phi['type'], phi['content']) for phi in detected}
    
    tp = len(gt_set & det_set)  # True Positives
    fp = len(det_set - gt_set)  # False Positives
    fn = len(gt_set - det_set)  # False Negatives
    
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
    
    return {
        'TP': tp,
        'FP': fp,
        'FN': fn,
        'Precision': precision,
        'Recall': recall,
        'F1': f1
    }

if __name__ == "__main__":
    # 範例使用
    test_text = "Patient 【PHI:NAME:P001】陳老先生【/PHI】, age 【PHI:AGE_OVER_89:A001】94【/PHI】"
    
    phi_list = parse_phi_tags(test_text)
    print("Parsed PHI:")
    for phi in phi_list:
        print(f"  - Type: {phi['type']}, Content: {phi['content']}, ID: {phi['id']}")
    
    # 載入測試文件
    df = pd.read_excel("data/test/test_phi_tagged_cases.xlsx")
    print(f"\\nLoaded {len(df)} test cases")
    
    total_phi = 0
    for idx, row in df.iterrows():
        case_id = row['Case ID']
        # 合併所有文本列
        full_text = ' '.join([str(row[col]) for col in df.columns[1:5] if pd.notna(row[col])])
        phi_list = parse_phi_tags(full_text)
        total_phi += len(phi_list)
        print(f"{case_id}: {len(phi_list)} PHI instances")
    
    print(f"\\nTotal PHI across all cases: {total_phi}")
"""

    parser_file = os.path.join(output_dir, "phi_tag_parser.py")
    with open(parser_file, 'w', encoding='utf-8') as f:
        f.write(parser_script)

    print(f"\n✅ Also generated: {parser_file}")
    print("   - Python script to parse PHI tags")
    print("   - Calculate Precision, Recall, F1 Score")

if __name__ == "__main__":
    generate_phi_tagged_test()
