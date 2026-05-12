"""Create de-identified artifacts by patching source containers in place.

The PHI engine works on extracted text, but production downloads should be
format-aware. This module generates a safe output artifact while the original
upload still exists, then the raw upload can be purged.
"""

from __future__ import annotations

import csv
import re
import shutil
from contextlib import suppress
from dataclasses import dataclass, field
from pathlib import Path

TEXT_EXTENSIONS = {".txt", ".md", ".markdown", ".json", ".xml", ".html", ".htm"}
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class DeidentifiedArtifact:
    """A generated de-identified file ready for download."""

    path: Path
    filename: str
    media_type: str
    format_preserved: bool
    engine: str
    warnings: list[str] = field(default_factory=list)


def build_deidentified_artifact(
    source_path: Path,
    original_filename: str,
    replacements: list[tuple[str, str]],
    output_dir: Path,
) -> DeidentifiedArtifact | None:
    """Build a de-identified output artifact from the original source file.

    Returns ``None`` only when the source format has no safe artifact writer yet.
    Callers may still fall back to text/JSON result export for legacy formats.
    """
    source_path = Path(source_path)
    suffix = source_path.suffix.lower()
    stem = _safe_stem(original_filename or source_path.name)
    output_dir.mkdir(parents=True, exist_ok=True)

    if suffix == ".csv":
        output_path = output_dir / f"{source_path.stem}_deidentified.csv"
        _patch_csv(source_path, output_path, replacements)
        return DeidentifiedArtifact(
            path=output_path,
            filename=f"{stem}_deidentified.csv",
            media_type="text/csv; charset=utf-8",
            format_preserved=True,
            engine="csv-dialect-patch",
        )

    if suffix == ".xlsx":
        output_path = output_dir / f"{source_path.stem}_deidentified.xlsx"
        _patch_xlsx(source_path, output_path, replacements)
        return DeidentifiedArtifact(
            path=output_path,
            filename=f"{stem}_deidentified.xlsx",
            media_type=XLSX_MEDIA_TYPE,
            format_preserved=True,
            engine="openpyxl-container-patch",
        )

    if suffix == ".xls":
        output_path = output_dir / f"{source_path.stem}_deidentified.xlsx"
        _convert_xls_to_xlsx(source_path, output_path, replacements)
        return DeidentifiedArtifact(
            path=output_path,
            filename=f"{stem}_deidentified.xlsx",
            media_type=XLSX_MEDIA_TYPE,
            format_preserved=False,
            engine="pandas-xls-fallback",
            warnings=["Legacy .xls is converted to .xlsx; workbook styling may not be preserved."],
        )

    if suffix in TEXT_EXTENSIONS:
        output_path = output_dir / f"{source_path.stem}_deidentified{suffix}"
        _patch_text(source_path, output_path, replacements)
        return DeidentifiedArtifact(
            path=output_path,
            filename=f"{stem}_deidentified{suffix}",
            media_type=_text_media_type(suffix),
            format_preserved=True,
            engine="text-span-patch",
        )

    if suffix == ".docx":
        return _try_asset_aware_docx(source_path, original_filename, replacements, output_dir)

    return None


def _safe_stem(filename: str) -> str:
    stem = Path(filename).stem or "result"
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", stem).strip()
    return stem[:160] or "result"


def _apply_replacements(value: str, replacements: list[tuple[str, str]]) -> str:
    masked = value
    for original, replacement in replacements:
        if original:
            masked = masked.replace(original, replacement)
    return masked


def _detect_text_encoding(path: Path) -> str:
    sample = path.read_bytes()[:4096]
    for encoding in ("utf-8-sig", "utf-8", "cp950", "big5"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def _patch_text(source_path: Path, output_path: Path, replacements: list[tuple[str, str]]) -> None:
    encoding = _detect_text_encoding(source_path)
    content = source_path.read_text(encoding=encoding, errors="replace")
    output_path.write_text(_apply_replacements(content, replacements), encoding="utf-8")


def _patch_csv(source_path: Path, output_path: Path, replacements: list[tuple[str, str]]) -> None:
    encoding = _detect_text_encoding(source_path)
    sample = source_path.read_text(encoding=encoding, errors="replace")[:8192]
    dialect = csv.excel
    with suppress(csv.Error):
        dialect = csv.Sniffer().sniff(sample)

    with (
        source_path.open("r", encoding=encoding, errors="replace", newline="") as src,
        output_path.open("w", encoding="utf-8-sig", newline="") as dst,
    ):
        reader = csv.reader(src, dialect)
        writer = csv.writer(
            dst,
            dialect=dialect,
            lineterminator=getattr(dialect, "lineterminator", "\n") or "\n",
        )
        for row in reader:
            writer.writerow([_apply_replacements(str(cell), replacements) for cell in row])


def _patch_xlsx(source_path: Path, output_path: Path, replacements: list[tuple[str, str]]) -> None:
    from openpyxl import load_workbook

    shutil.copy2(source_path, output_path)
    workbook = load_workbook(output_path)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or value.startswith("="):
                    continue
                masked = _apply_replacements(value, replacements)
                if masked != value:
                    cell.value = masked
    workbook.save(output_path)


def _convert_xls_to_xlsx(
    source_path: Path,
    output_path: Path,
    replacements: list[tuple[str, str]],
) -> None:
    import pandas as pd

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        workbook = pd.ExcelFile(source_path)
        for sheet_name in workbook.sheet_names:
            dataframe = pd.read_excel(
                workbook,
                sheet_name=sheet_name,
                dtype=str,
                keep_default_na=False,
            )
            masked = dataframe.map(lambda value: _apply_replacements(str(value), replacements))
            masked.to_excel(writer, sheet_name=str(sheet_name)[:31] or "Sheet1", index=False)


def _text_media_type(suffix: str) -> str:
    if suffix == ".json":
        return "application/json; charset=utf-8"
    if suffix in {".html", ".htm"}:
        return "text/html; charset=utf-8"
    if suffix == ".xml":
        return "application/xml; charset=utf-8"
    return "text/plain; charset=utf-8"


def _try_asset_aware_docx(
    source_path: Path,
    original_filename: str,
    replacements: list[tuple[str, str]],
    output_dir: Path,
) -> DeidentifiedArtifact | None:
    """Use asset-aware-mcp DFM when installed.

    The upstream package currently exposes its implementation through a generic
    top-level ``src`` package, so imports stay isolated here instead of leaking
    into the rest of the codebase.
    """
    try:
        artifact = _patch_docx_with_asset_aware(
            source_path,
            original_filename,
            replacements,
            output_dir,
        )
    except (ImportError, ModuleNotFoundError):
        return None
    return artifact


def _patch_docx_with_asset_aware(
    source_path: Path,
    original_filename: str,
    replacements: list[tuple[str, str]],
    output_dir: Path,
) -> DeidentifiedArtifact:
    from src.infrastructure.docx_adapter import DocxAdapter  # type: ignore[import-not-found]

    adapter = DocxAdapter()
    work_dir = output_dir / f"{source_path.stem}_dfm"
    ir = adapter.parse_to_ir(source_path, work_dir)
    changed_block_ids: set[str] = set()

    for block in ir.editable_blocks:
        original_content = block.content
        block.content = _apply_replacements(block.content, replacements)
        if block.runs:
            for run in block.runs:
                run.text = _apply_replacements(run.text, replacements)
        if block.content != original_content:
            changed_block_ids.add(block.id)

    output_path = output_dir / f"{source_path.stem}_deidentified.docx"
    adapter.ir_to_docx(ir, work_dir, output_path, changed_block_ids=changed_block_ids or None)
    return DeidentifiedArtifact(
        path=output_path,
        filename=f"{_safe_stem(original_filename)}_deidentified.docx",
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        format_preserved=True,
        engine="asset-aware-dfm",
        warnings=[] if changed_block_ids else ["DOCX contained no directly editable PHI spans."],
    )
